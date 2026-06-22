"""
Build a scalable SOIC ranking workbook from the local SOIC template and Chartink.

The generated workbook keeps the SOIC scoring logic formula-driven, but changes the
layout to one company per row so the full Chartink universe can be ranked.
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

import time
import numpy as np
import pandas as pd
import yfinance as yf
try:
    yf.set_tz_cache_location("logs/yfinance_cache")
except Exception:
    pass
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import sqlite3



import os
def load_dotenv():
    """Load variables from .env file into os.environ if it exists."""
    from pathlib import Path
    env_path = Path(".env")
    if env_path.exists():
        with open(env_path, "r") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    os.environ[key.strip()] = val.strip().strip('"').strip("'")

load_dotenv()


DEFAULT_TEMPLATE = (
    Path("SOIC Ranking Sheet.xlsx")
    if Path("SOIC Ranking Sheet.xlsx").exists()
    else Path("/Users/monitranjan/Downloads/SOIC Ranking Sheet.xlsx")
)
DEFAULT_OUTPUT = Path("outputs")

SCAN_CLAUSE = (
    "( {cash} ( "
    " daily close > daily ema ( daily close , 200 ) "
    "and daily close > ( daily max ( 252 , daily high ) * 0.80 ) "
    "and daily rsi(14) > 55 "
    "and market cap >= 500 "
    ") )"
)

METADATA_COLUMNS = (
    "sector as 'Sector', "
    "industry as 'Industry', "
    "marketcapname as 'Marketcap Name', "
    "market cap as 'Market Cap'"
)


@dataclass
class Criterion:
    section: str
    label: str
    question: str
    source_row: int
    score_formula: str
    validation_formula: str | None


def clean_name(name: str) -> str:
    """Clean company names for robust comparison."""
    if not name:
        return ""
    name = str(name).lower()
    # remove text in parentheses (e.g. "BSE (Bombay stock exchange) Ltd" -> "BSE Ltd")
    name = re.sub(r'\(.*?\)', '', name)
    # remove non-alphanumeric characters
    name = re.sub(r'[^a-z0-9\s]', '', name)
    words = name.split()
    # remove common corporate / industry suffixes
    ignore = {
        'limited', 'ltd', 'co', 'inc', 'corp', 'corporation', 
        'industries', 'industry', 'ind', 'pharma', 'pharmaceuticals', 
        'controls', 'chemicals', 'chemical', 'limitedlimit', 'limit', 
        'india', 'intl', 'international', 'sfb', 'bank', 'finance', 
        'financial', 'services', 'holding', 'holdings', 'investment', 
        'investments', 'capital', 'group', 'power', 'energy', 'infrastructure'
    }
    words = [w for w in words if w not in ignore]
    return ' '.join(words).strip()


def find_matching_company(
    target_name: str, existing_companies: Iterable[str]
) -> str | None:
    """Find a matching company name using cleaned fuzzy comparison."""
    target_clean = clean_name(target_name)
    if not target_clean:
        return None
    for ec in existing_companies:
        ec_clean = clean_name(ec)
        if not ec_clean:
            continue
        # Substring matching or exact match after cleaning
        if target_clean in ec_clean or ec_clean in target_clean:
            return ec
        # Check if first word matches and is significant (length >= 3)
        t_words = target_clean.split()
        ec_words = ec_clean.split()
        if t_words and ec_words:
            if len(t_words[0]) >= 3 and t_words[0] in ec_words:
                return ec
            if len(ec_words[0]) >= 3 and ec_words[0] in t_words:
                return ec
    return None


def extract_existing_company_inputs(
    template_path: Path, sheet_name: str, criteria: list[Criterion]
) -> dict[str, dict[str, str]]:
    """
    Extract existing qualitative inputs for companies in the template.
    Returns a dict mapping company name -> {criterion_label: input_value}.
    """
    if not template_path.exists():
        return {}
    
    # Read with data_only=True to get final text values rather than formulas
    wb = load_workbook(template_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    
    company_inputs: dict[str, dict[str, str]] = {}
    
    # Template company columns start at col 4 (D) and are every 2nd column
    for col in range(4, ws.max_column + 1, 2):
        company_name = ws.cell(7, col).value
        # Skip empty or placeholder/numerical columns
        if not company_name or isinstance(company_name, (int, float)) or "company" in str(company_name).lower() or str(company_name) == "0.0":
            continue
            
        company_name_str = str(company_name).strip()
        inputs = {}
        for c in criteria:
            val = ws.cell(c.source_row, col).value
            if val is not None:
                inputs[c.label] = str(val).strip()
                
        company_inputs[company_name_str] = inputs
        
    return company_inputs


def calculate_rsi(prices: pd.Series, period: int = 14) -> float:
    if prices is None or len(prices) < period + 1:
        return 50.0
    delta = prices.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    # Wilder's Smoothing (RMA) is an EMA with alpha = 1 / period
    avg_gain = gain.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    val = rsi.iloc[-1]
    return 50.0 if pd.isna(val) else float(val)


def calculate_adx(high: pd.Series, low: pd.Series, close: pd.Series, period: int = 14) -> float:
    if close is None or len(close) < period * 2:
        return 25.0
    try:
        tr1 = high - low
        tr2 = (high - close.shift(1)).abs()
        tr3 = (low - close.shift(1)).abs()
        tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
        
        up_move = high - high.shift(1)
        down_move = low.shift(1) - low
        
        plus_dm = np.where((up_move > down_move) & (up_move > 0), up_move, 0.0)
        minus_dm = np.where((down_move > up_move) & (down_move > 0), down_move, 0.0)
        
        # Smooth TR and DMs using Wilder's Smoothing (RMA)
        tr_smoothed = tr.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
        plus_dm_smoothed = pd.Series(plus_dm, index=tr.index).ewm(alpha=1/period, min_periods=period, adjust=False).mean()
        minus_dm_smoothed = pd.Series(minus_dm, index=tr.index).ewm(alpha=1/period, min_periods=period, adjust=False).mean()
        
        plus_di = 100 * (plus_dm_smoothed / tr_smoothed.replace(0, np.nan))
        minus_di = 100 * (minus_dm_smoothed / tr_smoothed.replace(0, np.nan))
        
        dx = 100 * (plus_di - minus_di).abs() / (plus_di + minus_di).replace(0, np.nan)
        # Smooth DX using a Simple Moving Average (SMA) to perfectly match TradingView's ADX and DI v4 indicator!
        adx = dx.rolling(window=period, min_periods=period).mean()
        val = adx.iloc[-1]
        return 25.0 if pd.isna(val) else float(val)
    except Exception:
        return 25.0


def calculate_vstop_signal(high: pd.Series, low: pd.Series, close: pd.Series, period: int = 14, multiplier: float = 3.0) -> str:
    """
    Calculate Volatility Trailing Stop (V-stop) signal.
    Returns 'Yes' if close is above the V-stop trailing line (bullish), 'No' if below (bearish).
    """
    if close is None or len(close) < period + 1:
        return "Yes"
    try:
        tr1 = high - low
        tr2 = (high - close.shift(1)).abs()
        tr3 = (low - close.shift(1)).abs()
        tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
        
        # Calculate ATR using Wilder's Smoothing (RMA)
        atr = tr.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
        
        highest_close = close.rolling(period).max()
        trailing_stop = highest_close - multiplier * atr
        
        curr_close = close.iloc[-1]
        curr_stop = trailing_stop.iloc[-1]
        
        if pd.isna(curr_close) or pd.isna(curr_stop):
            return "Yes"
            
        return "Yes" if curr_close >= curr_stop else "No"
    except Exception:
        return "Yes"


def calculate_vstop_price(high: pd.Series, low: pd.Series, close: pd.Series, period: int = 14, multiplier: float = 3.0) -> float:
    """Calculate the actual Volatility Trailing Stop (V-stop) line price."""
    if close is None or len(close) < period + 1:
        return 0.0
    try:
        tr1 = high - low
        tr2 = (high - close.shift(1)).abs()
        tr3 = (low - close.shift(1)).abs()
        tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
        # Calculate ATR using Wilder's Smoothing (RMA)
        atr = tr.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
        highest_close = close.rolling(period).max()
        trailing_stop = highest_close - multiplier * atr
        val = trailing_stop.iloc[-1]
        return 0.0 if pd.isna(val) else float(val)
    except Exception:
        return 0.0


def calculate_near_term_trigger_val(close: pd.Series, volume: pd.Series) -> str:
    """Determine the actionable catalyst reason for near-term triggers."""
    if close is None or len(close) < 20:
        return "Momentum Support"
    try:
        curr_price = close.iloc[-1]
        high_52w = close.rolling(52, min_periods=20).max().iloc[-1]
        near_high = curr_price >= 0.95 * high_52w
        
        if volume is not None and len(volume) >= 10:
            curr_vol = volume.iloc[-1]
            avg_vol = volume.rolling(10).mean().iloc[-1]
            vol_spike = curr_vol > 1.5 * avg_vol
        else:
            vol_spike = False
            
        if near_high and vol_spike:
            return "Volume Spike & Near 52W High Breakout"
        elif near_high:
            return "Near 52W High Breakout"
        elif vol_spike:
            return "Volume Spike / Accumulation"
        else:
            return "Earnings / Momentum Support"
    except Exception:
        return "Momentum Support"


def evaluate_excel_if_formula(formula: str, input_value: str) -> float:
    """
    Evaluate a nested Excel IF formula for a given input value.
    Example: '=IF(D8="Growing Industry",10,if(D8="Cyclical in an Upturn",5,0))'
    """
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return 0.0
    if input_value is None:
        return 0.0
    
    formula_lower = formula.lower()
    input_value_lower = str(input_value).lower().strip()
    
    # Match condition-value pairs from formula: "expected_value", score
    matches = re.findall(r'\"([^\"]+)\"\s*,\s*(-?\d+)', formula_lower)
    for expected_val, score_str in matches:
        if expected_val.strip() == input_value_lower:
            return float(score_str)
            
    # Check fallback value at the end of the nested IF formula
    fallback_match = re.search(r',([-\d]+)\s*\)+$', formula_lower)
    if fallback_match:
        return float(fallback_match.group(1))
        
    return 0.0


def clean_excel_formula(formula: str) -> str:
    """Clean Excel formula to ensure standard function names are strictly UPPERCASE."""
    if not formula or not isinstance(formula, str):
        return formula
    # Uppercase function keywords followed by '('
    def repl(m):
        return m.group(1).upper() + m.group(2)
    cleaned = re.sub(r'\b(if|sum|rank\.eq|rank|and|or|not|average)\b(\s*\()', repl, formula, flags=re.IGNORECASE)
    # Also replace RANK.EQ with RANK for openpyxl compatibility!
    cleaned = re.sub(r'\bRANK\.EQ\b', 'RANK', cleaned, flags=re.IGNORECASE)
    return cleaned


def get_automated_input_value(
    label: str, row: pd.Series, details: dict, is_bank: bool
) -> str | None:
    """Determine best automated qualitative or fundamental value based on internet data."""
    info = details.get("info", {})
    close_prices = details.get("close", None)
    high_prices = details.get("high", None)
    low_prices = details.get("low", None)

    sector = str(row.get("sector") or info.get("sector") or "").lower()
    industry = str(row.get("industry") or info.get("industry") or "").lower()
    company = str(row.get("company") or info.get("longName") or "").lower()
    symbol = str(row.get("symbol") or info.get("symbol") or "").upper()

    # 1. TECHNICAL INDICATORS
    if label == "RSI":
        rsi = calculate_rsi(close_prices)
        if rsi > 75:
            return "Above 75"
        elif rsi >= 50:
            return "Between 50 to 75"
        else:
            return "Below 45"
            
    if label == "ADX":
        adx = calculate_adx(high_prices, low_prices, close_prices)
        if adx > 40:
            return "Above 40"
        elif adx >= 20:
            return "20 to 40"
        else:
            return "0 to 20"

    if label == "Stage Analysis":
        if close_prices is not None and len(close_prices) >= 30:
            # 30-week EMA (Stan Weinstein Stage Analysis using Exponential Moving Average)
            ema30 = close_prices.ewm(span=30, min_periods=30, adjust=False).mean()
            curr_close = close_prices.iloc[-1]
            curr_ema = ema30.iloc[-1]
            prev_ema = ema30.iloc[-5] # 5 weeks ago
            slope = (curr_ema - prev_ema) / prev_ema
            if curr_close > curr_ema:
                return "Stage 2" if slope > 0.005 else "Stage 1"
            else:
                return "Stage 4" if slope < -0.005 else "Stage 3"
        return "Stage 2"

    if label in ("Primary Trend", "Secondary Trend"):
        return "Uptrend"

    if label == "Price Volume Action":
        if close_prices is not None and len(close_prices) >= 5:
            ret = (close_prices.iloc[-1] - close_prices.iloc[-5]) / close_prices.iloc[-5]
            return "Positive Trend" if ret >= 0 else "Negative Trend"
        return "Positive Trend"

    if label == "Resistance":
        if close_prices is not None and len(close_prices) >= 20:
            high_52w = close_prices.rolling(252, min_periods=20).max().iloc[-1]
            curr = close_prices.iloc[-1]
            if pd.notna(high_52w) and curr >= high_52w * 0.95:
                return "Blue sky scenario"
        return "Close to 52 week high or retracement"

    if label == "V-stop":
        return calculate_vstop_signal(high_prices, low_prices, close_prices)

    # 2. FUNDAMENTALS FROM YFINANCE INFO
    if label == "PAT Growth":
        growth = info.get("earningsGrowth")
        if growth is not None:
            growth = float(growth)
            if growth >= 0.20:
                return "More than 20%"
            elif growth >= 0.12:
                return "Between 12 to 20%"
            else:
                return "Less than 12%"
        return "More than 20%"  # default optimistic for momentum stock

    if label == "PE Ratio":
        pe = info.get("trailingPE")
        if pe is not None:
            return "Yes" if float(pe) > 40 else "No"
        return "No"

    if label == "PEG Ratio":
        peg = info.get("trailingPegRatio")
        if peg is None or pd.isna(peg):
            # Robust manual fallback calculation
            pe = info.get("trailingPE") or info.get("forwardPE")
            growth = info.get("earningsGrowth") or info.get("revenueGrowth")
            if pe is not None and growth is not None:
                pe = float(pe)
                growth = float(growth)
                if growth > 0:
                    pct_growth = growth if growth > 1.0 else growth * 100.0
                    peg = pe / pct_growth
                else:
                    peg = 99.0  # Failsafe for negative/flat growth to place in "Above 2"
        if peg is not None:
            peg = float(peg)
            if peg < 1.5:
                return "Less than 1.5"
            elif peg <= 2.0:
                return "Between 1.5 to 2"
            else:
                return "Above 2"
        return "Less than 1.5"

    if label == "CFO to EBITDA":
        cfo = info.get("operatingCashflow")
        ebitda = info.get("ebitda")
        if cfo is not None and ebitda is not None and float(ebitda) != 0:
            ratio = float(cfo) / float(ebitda)
            if ratio <= 0:
                return "Negative"
            elif ratio >= 0.70:
                return "Above 70%"
            elif ratio >= 0.50:
                return "Between 50% to 70%"
            else:
                debt_eq = info.get("debtToEquity", 0)
                if debt_eq and float(debt_eq) > 100:
                    return "Less than 50% and Borrowings are increasing"
                return "Less than 50%"
        return "Above 70%"

    if label == "Debt to Equity":
        de = info.get("debtToEquity")
        if de is not None:
            ratio = float(de) / 100.0  # de in yfinance is typically % (e.g. 3.79 -> 0.038)
            if ratio > 1.2:
                return "More than 1.2"
            elif ratio >= 0.7:
                return "Between 0.7 to 1.2"
            else:
                return "Less than 0.7"
        return "Less than 0.7"

    if label == "Stock History":
        hist_len = details.get("history_len", 0)
        return "Yes" if hist_len > 200 else "No"

    if label == "Auditor Resignation":
        return "No"

    if label == "Contingent Liability":
        return "Less than 15%"

    if label == "Pledging":
        return "No"

    # Specific Bank Price to Book
    if label == "Price to Book Value" and is_bank:
        pb = info.get("priceToBook")
        growth = info.get("earningsGrowth") or info.get("revenueGrowth") or 0.15
        if pb is not None:
            pb = float(pb)
            growth = float(growth)
            if pb < 2.5:
                return "P/BV less than 2.5x and Growth more than 20%" if growth > 0.20 else "P/BV less than 2.5x and Growth less than 20%"
            elif pb <= 5.0:
                return "P/BV between 2.5x to 5x and Growth more than 20%" if growth > 0.20 else "P/BV between 2.5x to 5x and Growth less than 20%"
            else:
                return "P/BV above 5x"
        return "P/BV less than 2.5x and Growth more than 20%"

    # 3. SECTOR/INDUSTRY QUALITATIVE MATRICES
    if label == "Industry Growth":
        if any(w in sector or w in industry for w in ["health", "pharm", "tech", "fmcg", "consumer defensive", "retail"]):
            return "Growing Industry"
        if any(w in sector or w in industry for w in ["basic materials", "energy", "industrials", "metal", "mining"]):
            return "Cyclical in an Upturn"
        return "Growing Industry"

    if label == "Competitiveness":
        if "vbl" in symbol or "varun" in company:
            return "Duopoly That Cooperate" if not is_bank else "Fragmented with High ROE"
        if "bse" in symbol:
            return "Duopoly That Cooperate" if not is_bank else "Fragmented with High ROE"
        if any(w in sector or w in industry for w in ["utility", "energy", "telecom"]):
            return "Oligopoly" if not is_bank else "Fragmented with High ROE"
        return "Too Many Players with High ROCE" if not is_bank else "Fragmented with High ROE"

    if label == "Regulation":
        if any(w in sector or w in industry for w in ["health", "pharm", "financial", "bank", "utility", "energy", "telecom"]):
            return "Yes"
        return "No"

    if label == "Sector Headwinds/Tailwind":
        if close_prices is not None and len(close_prices) >= 26:
            # 6-month momentum (26 weeks)
            ret_6m = (close_prices.iloc[-1] - close_prices.iloc[-26]) / close_prices.iloc[-26]
            return "Tailwind" if ret_6m >= 0 else "Headwind"
        return "Tailwind"

    if label == "Industry Type":
        if any(w in sector or w in industry for w in ["basic materials", "energy", "metal", "commodit"]):
            return "Cyclical" if not is_bank else "Highly Cyclical"
        return "Structural (10 Years Consistency)"

    if label == "Margins Stability":
        if any(w in sector or w in industry for w in ["basic materials", "energy", "metal", "commodit"]):
            return "Volatile"
        return "Stable"

    if label == "Business Type":
        if any(w in sector or w in industry for w in ["consumer cyclical", "fmcg", "retail"]):
            return "B2C"
        return "B2B"

    if label == "Type of Revenue":
        if any(w in sector or w in industry for w in ["tech", "soft", "fmcg", "health"]):
            return "Higher Annuity, Lower One Time"
        return "Higher One Time, Lower Annuity"

    if label == "Margins":
        return "No"

    if label == "Growth":
        return "Yes"

    if label == "Near term triggers":
        if close_prices is not None and len(close_prices) >= 20:
            curr_price = close_prices.iloc[-1]
            high_52w = close_prices.rolling(52, min_periods=20).max().iloc[-1]
            near_high = curr_price >= 0.95 * high_52w
            
            volume_prices = details.get("volume", None)
            if volume_prices is not None and len(volume_prices) >= 10:
                curr_vol = volume_prices.iloc[-1]
                avg_vol = volume_prices.rolling(10).mean().iloc[-1]
                vol_spike = curr_vol > 1.5 * avg_vol
            else:
                vol_spike = False
                
            return "Yes" if (near_high or vol_spike) else "No"
        return "Yes"

    if label == "Capital Allocation":
        return "Good"

    if label == "Medium term Growth":
        return "Yes for more than 2 years"

    # Specific Bank fields
    if is_bank:
        if label in ("NIM", "CASA Ratio", "CAR Ratio", "Return on Assets", "AUM Growth"):
            return "Yes"
        if label in ("Return on Equity", "High P/BV"):
            return "Yes"
        if label in ("Asset Quality Divergence", "Reported loss", "CEO Resignation", "GNPA", "KMP Resignation"):
            return "No"

    if label == "Scan Matches":
        scans = row.get("Scans") or 0
        return str(scans)

    return None


def map_stockscans_index_to_yfinance(index_id: str, index_name: str) -> tuple[str, str] | None:
    """
    Map a StockScans index ID/Name to a supported Yahoo Finance index ticker and name.
    Returns: (yf_ticker, yf_name) or None if no map is found.
    """
    id_upper = str(index_id or "").upper()
    name_lower = str(index_name or "").lower()
    
    # Direct Ticker Mapping for known indices
    if "CNXFMCG" in id_upper or "fmcg" in name_lower:
        return "^CNXFMCG", "Nifty FMCG"
    if "CNXPHARMA" in id_upper or "pharma" in name_lower or "healthcare" in name_lower:
        return "^CNXPHARMA", "Nifty Pharma"
    if "CNXIT" in id_upper or "it" in name_lower:
        return "^CNXIT", "Nifty IT"
    if "CNXAUTO" in id_upper or "auto" in name_lower:
        return "^CNXAUTO", "Nifty Auto"
    if "CNXMETAL" in id_upper or "metal" in name_lower:
        return "^CNXMETAL", "Nifty Metal"
    if "CNXREALTY" in id_upper or "realty" in name_lower:
        return "^CNXREALTY", "Nifty Realty"
    if "CNXENERGY" in id_upper or "energy" in name_lower:
        return "^CNXENERGY", "Nifty Energy"
    if "CNXINFRA" in id_upper or "infra" in name_lower or "indust" in name_lower:
        return "^CNXINFRA", "Nifty Infra"
    if "BANK" in id_upper or "bank" in name_lower or "finance" in name_lower:
        if "PSU" in id_upper or "psu" in name_lower:
            return "^CNXPSUBANK", "Nifty PSU Bank"
        return "^NSEBANK", "Nifty Bank"
    if "MEDIA" in id_upper or "media" in name_lower:
        return "^CNXMEDIA", "Nifty Media"
        
    # General Market Indices
    if "CNX500" in id_upper or "500" in name_lower:
        return "^CRSLDX", "Nifty 500"
    if "CNX200" in id_upper or "200" in name_lower:
        return "^CNX200", "Nifty 200"
    if "CNX100" in id_upper or "100" in name_lower:
        return "^CNX100", "Nifty 100"
    if "NIFTY50" in id_upper or "nifty 50" in name_lower or "nifty" in name_lower or "sensex" in name_lower:
        return "^NSEI", "Nifty 50"
        
    return None


def fetch_all_stockscans_indices(symbols: list[str]) -> dict[str, tuple[str, str]]:
    """Fetch indices list for all symbols from StockScans API in parallel."""
    cookie = os.environ.get("STOCKSCANS_COOKIE", "")
    headers = {
        "accept": "application/json",
        "cookie": cookie,
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, Gecko) Chrome/148.0.0.0 Safari/537.36"
    }
    
    print(f"Fetching StockScans indices data for {len(symbols)} stocks in parallel...")
    results = {}
    
    def fetch_one(symbol: str):
        for exchange in ["NSE", "BSE"]:
            url = f"https://www.stockscans.in/api/company/indices/{exchange}:{symbol}"
            try:
                r = requests.get(url, headers=headers, timeout=10)
                if r.status_code == 200:
                    data = r.json()
                    indices = data.get("indices", [])
                    if indices:
                        for idx_item in indices:
                            mapped = map_stockscans_index_to_yfinance(idx_item.get("companyId"), idx_item.get("Name"))
                            if mapped:
                                return symbol, mapped
                        first_idx = indices[0]
                        return symbol, (first_idx.get("companyId"), first_idx.get("Name"))
            except Exception:
                continue
        return symbol, None

    from concurrent.futures import ThreadPoolExecutor, as_completed
    with ThreadPoolExecutor(max_workers=30) as executor:
        futures = {executor.submit(fetch_one, sym): sym for sym in symbols}
        for future in as_completed(futures):
            sym = futures[future]
            try:
                symbol, mapped = future.result()
                if mapped:
                    results[symbol] = mapped
            except Exception as e:
                print(f"⚠️ Error fetching StockScans indices for {sym}: {e}")
                
    return results


def get_benchmark_ticker(sector: str, industry: str) -> tuple[str, str] | None:
    """
    Map stock sector and industry to the most appropriate Nifty benchmark index.
    Returns: (benchmark_ticker, benchmark_name) or None if not matching standard sectoral index.
    """
    s = str(sector or "").strip().lower()
    ind = str(industry or "").strip().lower()
    
    # Healthcare / Pharma
    if "pharma" in s or "pharma" in ind or "healthcare" in s or "hospital" in ind:
        return "^CNXPHARMA", "Nifty Pharma"
    
    # Banking & Financial Services
    if "psu bank" in ind or "psu bank" in s:
        return "^CNXPSUBANK", "Nifty PSU Bank"
    if "bank" in s or "bank" in ind or "financial" in s or "financial" in ind or "nbfc" in s or "nbfc" in ind:
        return "^NSEBANK", "Nifty Bank"
        
    # Information Technology
    if "it" in s or "i.t" in s or "software" in ind or "data centre" in ind:
        return "^CNXIT", "Nifty IT"
        
    # Auto & Auto Ancillaries
    if "auto" in s or "auto" in ind:
        return "^CNXAUTO", "Nifty Auto"
        
    # Metals & Mining
    if "metal" in s or "mining" in s or "metal" in ind or "mining" in ind or "steel" in ind or "aluminium" in ind:
        return "^CNXMETAL", "Nifty Metal"
        
    # Realty
    if "realty" in s or "realty" in ind or "real estate" in s or "real estate" in ind:
        return "^CNXREALTY", "Nifty Realty"
        
    # FMCG / Consumer Discretionary (FMCG proxy)
    if "fmcg" in s or "fmcg" in ind or "personal care" in ind or "agricultural" in ind:
        return "^CNXFMCG", "Nifty FMCG"
        
    # Energy & Utilities
    if "energy" in s or "power" in s or "utilities" in s or "power" in ind or "oil" in ind:
        return "^CNXENERGY", "Nifty Energy"
        
    # Infrastructure & Construction
    if "infra" in s or "infrastructure" in s or "construction" in ind or "building materials" in s or "cables" in ind:
        return "^CNXINFRA", "Nifty Infra"
        
    # Media
    if "media" in s or "media" in ind:
        return "^CNXMEDIA", "Nifty Media"
        
    return None


def fetch_benchmark_indices() -> dict[str, dict]:
    """Fetch 2-year weekly history for Nifty benchmark indices."""
    benchmarks = [
        "^NSEI", "^CRSLDX", "^NSEBANK", "^CNXIT", "^CNXPHARMA", 
        "^CNXFMCG", "^CNXAUTO", "^CNXMETAL", "^CNXREALTY", 
        "^CNXENERGY", "^CNXINFRA", "^CNXMEDIA", "^CNXPSUBANK"
    ]
    print(f"Fetching weekly data for {len(benchmarks)} benchmark indices in parallel...")
    results = {}
    
    def fetch_one(ticker_symbol: str):
        for attempt in range(3):
            try:
                ticker = yf.Ticker(ticker_symbol)
                hist = ticker.history(period="2y", interval="1wk")
                if hist is not None and not hist.empty:
                    return ticker_symbol, {
                        "close": hist["Close"].squeeze(),
                        "high": hist["High"].squeeze(),
                        "low": hist["Low"].squeeze(),
                        "volume": hist["Volume"].squeeze(),
                    }
            except Exception as e:
                time.sleep(0.5)
        print(f"⚠️ Error fetching benchmark index {ticker_symbol}")
        return ticker_symbol, {}

    from concurrent.futures import ThreadPoolExecutor, as_completed
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_one, t): t for t in benchmarks}
        for future in as_completed(futures):
            t, data = future.result()
            if data:
                results[t] = data
                
    return results


def calculate_oneil_relative_strength(
    stock_series: pd.Series, 
    bench_series: pd.Series
) -> tuple[float | None, float | None, float | None, float | None, float | None]:
    """
    Calculate date-aligned 1Y stock return, benchmark return, TradingView RS Spread, RS Line,
    and O'Neil Weighted Score (2*q4 + q3 + q2 + q1)/5 based on weekly series.
    Returns: (stock_1y_ret, bench_1y_ret, rs_spread, rs_line_now, weighted_score)
    """
    if stock_series is None or len(stock_series) < 2 or bench_series is None or len(bench_series) < 2:
        return None, None, None, None, None
        
    # Align by intersecting the date indices
    common = stock_series.index.intersection(bench_series.index)
    if len(common) < 2:
        return None, None, None, None, None
        
    s = stock_series.loc[common]
    b = bench_series.loc[common]
    
    # Restrict to last 53 weeks (approx 1 year of weekly close prices)
    if len(s) > 53:
        s = s.iloc[-53:]
        b = b.iloc[-53:]
        
    if len(s) < 2:
        return None, None, None, None, None
        
    # Standard 1Y Returns
    s_ret = (s.iloc[-1] / s.iloc[0] - 1) * 100.0
    b_ret = (b.iloc[-1] / b.iloc[0] - 1) * 100.0
    
    # TradingView comparative relative strength formula
    rs_spread = ((s.iloc[-1] / s.iloc[0]) / (b.iloc[-1] / b.iloc[0]) - 1) * 100.0
    
    # RS Line (stock relative performance trend)
    rs_line_series = (s / s.iloc[0]) / (b / b.iloc[0])
    rs_line_now = rs_line_series.iloc[-1]
    
    # O'Neil weighted score calculation
    q = len(s) // 4
    if q < 1:
        return round(s_ret, 2), round(b_ret, 2), round(rs_spread, 2), round(rs_line_now, 4), round(s_ret, 2)
        
    q1 = (s.iloc[q] / s.iloc[0] - 1) * 100.0
    q2 = (s.iloc[2 * q] / s.iloc[q] - 1) * 100.0
    q3 = (s.iloc[3 * q] / s.iloc[2 * q] - 1) * 100.0
    q4 = (s.iloc[-1] / s.iloc[3 * q] - 1) * 100.0
    weighted_score = (2 * q4 + q3 + q2 + q1) / 5.0
    
    return (
        round(s_ret, 2), 
        round(b_ret, 2), 
        round(rs_spread, 2), 
        round(rs_line_now, 4), 
        round(weighted_score, 2)
    )


def add_rs_rating_conditional_formatting(ws, col: int, start_row: int, end_row: int) -> None:
    """Add conditional formatting to RS Rating: >=95 green bold, >=90 light green, >=80 amber."""
    # Dark Green for >= 95
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100", bold=True)
    
    # Light Green for 90 to 94
    light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    light_green_font = Font(color="375623", bold=False)
    
    # Amber for 80 to 89
    amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    amber_font = Font(color="9C6500", bold=False)
    
    col_letter = get_column_letter(col)
    range_str = f"{col_letter}{start_row}:{col_letter}{end_row}"
    
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator='greaterThanOrEqual', formula=['95'], stopIfTrue=True, fill=green_fill, font=green_font)
    )
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator='between', formula=['90', '94'], stopIfTrue=True, fill=light_green_fill, font=light_green_font)
    )
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator='between', formula=['80', '89'], stopIfTrue=True, fill=amber_fill, font=amber_font)
    )


def calculate_global_rs_ratings(
    all_symbols: list[str], 
    universe: pd.DataFrame, 
    yfinance_data: dict[str, dict], 
    benchmark_data: dict[str, dict],
    stockscans_indices: dict[str, tuple[str, str]]
) -> dict[str, dict]:
    """
    Calculate O'Neil Weighted Score and RS Rating percentile (1-99) for all symbols.
    Returns a dict mapping symbol -> {
        "bench_name": bench_name,
        "stock_1y_ret": stock_1y_ret,
        "bench_1y_ret": bench_1y_ret,
        "rs_spread": rs_spread,
        "weighted_score": weighted_score,
        "rs_rating": rs_rating,
        "rs_status": rs_status
    }
    """
    scores_list = []
    rs_raw = {}
    
    for symbol in all_symbols:
        details = yfinance_data.get(symbol, {})
        if not details:
            continue
            
        # Sector benchmark mapping
        # Look up sector/industry in universe first
        univ_match = universe[universe["symbol"] == symbol]
        if not univ_match.empty:
            sector = univ_match.iloc[0].get("sector")
            industry = univ_match.iloc[0].get("industry")
        else:
            info = details.get("info", {})
            sector = info.get("sector", "")
            industry = info.get("industry", "")
            
        mapped_bench = get_benchmark_ticker(sector, industry)
        if mapped_bench is not None:
            bench_ticker, bench_name = mapped_bench
        else:
            api_mapped = stockscans_indices.get(symbol)
            if api_mapped is not None:
                bench_ticker, bench_name = api_mapped
            else:
                bench_ticker, bench_name = "^NSEI", "Nifty 50"
                
        if bench_ticker not in benchmark_data:
            bench_ticker, bench_name = "^NSEI", "Nifty 50"
            
        bench_series = benchmark_data.get(bench_ticker, {}).get("close")
        stock_series = details.get("close")
        
        stock_1y_ret, bench_1y_ret, rs_spread, rs_line_now, weighted_score = calculate_oneil_relative_strength(
            stock_series, bench_series
        )
        
        rs_raw[symbol] = {
            "bench_name": bench_name,
            "stock_1y_ret": stock_1y_ret,
            "bench_1y_ret": bench_1y_ret,
            "rs_spread": rs_spread,
            "weighted_score": weighted_score,
        }
        if weighted_score is not None:
            scores_list.append((symbol, weighted_score))
            
    # Calculate global percentile rank (1-99)
    valid_scores = sorted([score for sym, score in scores_list])
    total_valid = len(valid_scores)
    
    def get_percentile_rating(score) -> int | str:
        if score is None or total_valid == 0:
            return ""
        count = sum(1 for s in valid_scores if s <= score)
        percentile = (count / total_valid) * 99.0
        return max(1, min(99, int(round(percentile))))
        
    global_rs = {}
    for symbol, raw in rs_raw.items():
        score = raw["weighted_score"]
        rating = get_percentile_rating(score)
        rs_spread = raw["rs_spread"]
        status = "Outperforming" if (rs_spread is not None and rs_spread > 0) else "Underperforming" if rs_spread is not None else ""
        
        global_rs[symbol] = {
            "bench_name": raw["bench_name"],
            "stock_1y_ret": raw["stock_1y_ret"],
            "bench_1y_ret": raw["bench_1y_ret"],
            "rs_spread": raw["rs_spread"],
            "weighted_score": raw["weighted_score"],
            "rs_rating": rating,
            "rs_status": status
        }
        
    return global_rs


def fetch_single_stock_details(symbol: str) -> tuple[str, dict]:
    """Download history and info for a single stock."""
    # Gentle sleep to prevent hitting API limits
    time.sleep(0.02)
    for suffix in [".NS", ".BO"]:
        ticker_symbol = symbol + suffix
        try:
            ticker = yf.Ticker(ticker_symbol)
            hist = ticker.history(period="2y", interval="1wk")
            if hist is not None and len(hist) >= 30:
                info = ticker.info
                return symbol, {
                    "info": info,
                    "close": hist["Close"].squeeze(),
                    "high": hist["High"].squeeze(),
                    "low": hist["Low"].squeeze(),
                    "volume": hist["Volume"].squeeze(),
                    "history_len": len(hist),
                }
        except Exception:
            continue
    return symbol, {}


def fetch_all_stocks_details(symbols: list[str]) -> dict[str, dict]:
    """Fetch yfinance details for a list of stock symbols in parallel."""
    print(f"Fetching details for {len(symbols)} stocks in parallel (30 threads)...")
    results = {}
    from concurrent.futures import ThreadPoolExecutor, as_completed
    with ThreadPoolExecutor(max_workers=30) as executor:
        futures = {executor.submit(fetch_single_stock_details, sym): sym for sym in symbols}
        for i, future in enumerate(as_completed(futures), 1):
            sym = futures[future]
            try:
                symbol, data = future.result()
                if data:
                    results[symbol] = data
                    if i % 30 == 0 or i == len(symbols):
                        print(f"  [{i}/{len(symbols)}] Fetched details for {symbol}")
            except Exception as e:
                print(f"  [{i}/{len(symbols)}] Error fetching {sym}: {e}")
    return results


def fetch_chartink_universe() -> pd.DataFrame:
    """Fetch base Chartink rows and enrich them with sector/industry/market cap."""
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36"
            )
        }
    )
    r = session.get("https://chartink.com/screener", timeout=90)
    r.raise_for_status()
    csrf = BeautifulSoup(r.content, "html.parser").find(
        "meta", {"name": "csrf-token"}
    )["content"]
    headers = {
        "Referer": "https://chartink.com/screener",
        "x-csrf-token": csrf,
        "X-Requested-With": "XMLHttpRequest",
    }

    base = _post_chartink(session, headers, {"scan_clause": SCAN_CLAUSE})
    meta = _post_chartink(
        session,
        headers,
        {"scan_clause": SCAN_CLAUSE, "column_clause": METADATA_COLUMNS},
    )
    meta_cols = [
        "nsecode",
        "sector",
        "industry",
        "marketcap name",
        "market cap",
    ]
    enriched = base.merge(meta[meta_cols], on="nsecode", how="left")
    enriched = enriched.rename(
        columns={
            "nsecode": "symbol",
            "name": "company",
            "bsecode": "bse_code",
            "per_chg": "today_return_pct",
            "marketcap name": "marketcap_bucket",
            "market cap": "marketcap_cr",
        }
    )
    numeric_cols = ["close", "today_return_pct", "volume", "marketcap_cr"]
    for col in numeric_cols:
        enriched[col] = pd.to_numeric(enriched[col], errors="coerce")
    return enriched.sort_values("today_return_pct", ascending=False).reset_index(
        drop=True
    )


def _post_chartink(
    session: requests.Session, headers: dict[str, str], payload: dict[str, str]
) -> pd.DataFrame:
    resp = session.post(
        "https://chartink.com/screener/process",
        headers=headers,
        data=payload,
        timeout=120,
    )
    resp.raise_for_status()
    data = resp.json()
    if data.get("scan_error"):
        raise RuntimeError(f"Chartink scan error: {data['scan_error']}")
    return pd.DataFrame(data.get("data", []))


def extract_criteria(template_path: Path, sheet_name: str) -> list[Criterion]:
    wb = load_workbook(template_path, data_only=False)
    ws = wb[sheet_name]
    criteria: list[Criterion] = []
    section = ""
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 2).value
        formula = ws.cell(row, 5).value
        question = ws.cell(row, 3).value
        if label and not (isinstance(formula, str) and formula.startswith("=")):
            section = str(label)
        if not (isinstance(formula, str) and formula.startswith("=")):
            continue
        if not label or str(question or "").strip().lower() == "condition":
            continue
        criteria.append(
            Criterion(
                section=section,
                label=str(label),
                question=str(question or ""),
                source_row=row,
                score_formula=formula,
                validation_formula=find_validation_formula(ws, row),
            )
        )
    return criteria


def find_validation_formula(ws, row: int) -> str | None:
    cell = f"D{row}"
    for dv in ws.data_validations.dataValidation:
        if cell in dv:
            return dv.formula1
    return None


def is_financial_company(row: pd.Series) -> bool:
    text = " ".join(
        str(row.get(col, "")).lower()
        for col in ["symbol", "company", "sector", "industry"]
    )
    patterns = [
        "bank",
        "finance",
        "financial",
        "nbfc",
        "asset management",
        "housing finance",
        "capital services",
        "creditcare",
    ]
    return any(pattern in text for pattern in patterns)


def load_scanner_symbols() -> list[str]:
    """Read today's scanner outputs or fallback to the most recent one."""
    from datetime import date
    csv_path = Path("logs") / f"signals_{date.today()}.csv"
    if not csv_path.exists():
        csv_files = sorted(Path("logs").glob("signals_*.csv"))
        if csv_files:
            csv_path = csv_files[-1]
        else:
            return []
    try:
        df = pd.read_csv(csv_path)
        if "ticker" in df.columns:
            return df["ticker"].dropna().unique().tolist()
    except Exception:
        pass
    return []


def fetch_single_stockscans_details(symbol: str) -> tuple[str, dict]:
    """Fetch StockScans search-company data dynamically using your authtoken."""
    cookie = os.environ.get("STOCKSCANS_COOKIE", "")
    headers = {
        "accept": "application/json",
        "cookie": cookie,
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, Gecko) Chrome/148.0.0.0 Safari/537.36"
    }
    # Gentle wait to avoid rate limit
    time.sleep(0.01)
    for exchange in ["NSE", "BSE"]:
        url = f"https://www.stockscans.in/api/company/scans/search-company/{exchange}:{symbol}"
        try:
            r = requests.get(url, headers=headers, timeout=12)
            if r.status_code == 200:
                return symbol, r.json()
        except Exception:
            continue
    return symbol, {}


def fetch_all_stockscans_details(symbols: list[str]) -> dict[str, dict]:
    """Fetch StockScans search-company data in parallel using 30 threads."""
    print(f"Fetching StockScans search-company data for {len(symbols)} stocks in parallel...")
    results = {}
    from concurrent.futures import ThreadPoolExecutor, as_completed
    with ThreadPoolExecutor(max_workers=30) as executor:
        futures = {executor.submit(fetch_single_stockscans_details, sym): sym for sym in symbols}
        for i, future in enumerate(as_completed(futures), 1):
            sym = futures[future]
            try:
                symbol, data = future.result()
                if data:
                    results[symbol] = data
                    if i % 50 == 0 or i == len(symbols):
                        print(f"  [{i}/{len(symbols)}] Fetched StockScans details for {symbol}")
            except Exception as e:
                print(f"  [{i}/{len(symbols)}] Error fetching StockScans for {sym}: {e}")
    return results


STOCKSCANS_STATUS = {"status": "success", "message": "", "fetched_live": False}

def get_stockscans_common_stocks_data() -> dict:
    """Fetch live StockScans common-stocks (scan matches) list dynamically or fallback to local JSON file."""
    global STOCKSCANS_STATUS
    import json
    
    cookie = os.environ.get("STOCKSCANS_COOKIE", "")
    
    # 1. Attempt to fetch dynamically from API
    try:
        url = "https://www.stockscans.in/api/user/saved-scans/common-stocks"
        headers = {
            "accept": "application/json",
            "content-type": "application/json",
            "cookie": cookie,
            "origin": "https://www.stockscans.in",
            "referer": "https://www.stockscans.in/scan-match",
            "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, Gecko) Chrome/148.0.0.0 Safari/537.36"
        }
        payload = {"includePopular": True}
        print("Attempting to fetch live StockScans common stocks (scan matches)...")
        r = requests.post(url, headers=headers, json=payload, timeout=20)
        
        if r.status_code == 200:
            data = r.json()
            if "companies" in data and len(data["companies"]) > 0:
                print(f"Successfully fetched {len(data['companies'])} live StockScans companies!")
                STOCKSCANS_STATUS["status"] = "success"
                STOCKSCANS_STATUS["fetched_live"] = True
                STOCKSCANS_STATUS["message"] = "Fetched live successfully"
                
                # Cache locally in workspace for convenience and resilience
                try:
                    cache_paths = [
                        Path("scan_matched_data.json"),
                        Path("/Users/monitranjan/.gemini/antigravity/scratch/scan_matched_data.json")
                    ]
                    for cp in cache_paths:
                        try:
                            cp.parent.mkdir(parents=True, exist_ok=True)
                            with open(cp, "w") as f:
                                json.dump(data, f, indent=2)
                        except Exception:
                            pass
                except Exception:
                    pass
                return data
            else:
                print("Live response loaded but companies list was empty.")
                STOCKSCANS_STATUS["status"] = "empty_response"
                STOCKSCANS_STATUS["message"] = "Response returned empty companies list."
        else:
            print(f"Live API request failed with status code: {r.status_code}")
            if r.status_code in [401, 403]:
                STOCKSCANS_STATUS["status"] = "expired"
                STOCKSCANS_STATUS["message"] = f"StockScans session expired (HTTP {r.status_code}). Please update cookie."
            else:
                STOCKSCANS_STATUS["status"] = "failed"
                STOCKSCANS_STATUS["message"] = f"API returned error status code: {r.status_code}"
    except Exception as e:
        print(f"Error fetching live StockScans common stocks API: {e}")
        STOCKSCANS_STATUS["status"] = "failed"
        STOCKSCANS_STATUS["message"] = f"Connection error: {str(e)}"
        
    # 2. Fallback to reading the local cached file
    print("Falling back to local cached scan_matched_data.json file...")
    paths_to_try = [
        Path("scan_matched_data.json"),
        Path("/Users/monitranjan/.gemini/antigravity/scratch/scan_matched_data.json")
    ]
    
    for path in paths_to_try:
        if path.exists():
            try:
                with open(path, "r") as f:
                    data = json.load(f)
                if "companies" in data:
                    print(f"Successfully loaded {len(data.get('companies', []))} companies from fallback: {path}")
                    return data
            except Exception as e:
                print(f"Error reading fallback file {path}: {e}")
                
    print("WARNING: No StockScans data fetched or loaded from fallback caches. Returning empty structure.")
    return {}


CONFLUENCE_EMERGING_REPORT = ""
CONFLUENCE_EMERGING_HTML = ""




def generate_automated_reports(
    universe: pd.DataFrame,
    scanner_symbols: list[str],
    scan_matched_symbols: list[str],
    yfinance_data: dict[str, dict],
    scans_dict: dict[str, int],
    industry_dict: dict[str, str],
    signals_dict: dict[str, dict]
) -> None:
    """Generate Markdown and HTML reports for High-Conviction Confluences (Count=3) and Emerging Leaders."""
    global CONFLUENCE_EMERGING_REPORT, CONFLUENCE_EMERGING_HTML
    from datetime import date, datetime
    from pathlib import Path
    import json
    
    # Scan for existing research reports in outputs/reports/ to highlight compilation dates
    reports_dir = Path("outputs") / "reports"
    existing_reports = {}
    if reports_dir.exists():
        for filepath in reports_dir.glob("*_equity_report_*.md"):
            filename = filepath.name
            try:
                parts = filename.replace(".md", "").split("_equity_report_")
                if len(parts) == 2:
                    sym = parts[0]
                    date_part = parts[1]
                    dt = datetime.strptime(date_part, "%Y-%m-%d")
                    formatted_date = dt.strftime("%d-%b-%Y")
                    if sym not in existing_reports or dt > datetime.strptime(existing_reports[sym]["raw_date"], "%Y-%m-%d"):
                        existing_reports[sym] = {
                            "formatted": formatted_date,
                            "raw_date": date_part
                        }
            except Exception:
                pass
                
    today_str = date.today().strftime("%d %b %Y")
    
    # 1. Collect Common Count = 3 stocks
    set_universe = set(universe["symbol"].dropna().unique())
    set_scanner = set(scanner_symbols)
    set_scan_match = set(scan_matched_symbols)
    
    confluence_3_rows = []
    union_symbols = set_universe | set_scanner | set_scan_match
    for sym in union_symbols:
        in_univ = sym in set_universe
        in_scan = sym in set_scanner
        in_match = sym in set_scan_match
        
        if in_univ and in_scan and in_match:
            company_name = sym
            sector = "unknown"
            mcap_cr = 0.0
            close = 0.0
            ret1d = 0.0
            
            univ_match = universe[universe["symbol"] == sym]
            if not univ_match.empty:
                company_name = univ_match.iloc[0]["company"]
                sector = univ_match.iloc[0]["sector"]
                mcap_cr = univ_match.iloc[0].get("marketcap_cr", 0.0)
                close = univ_match.iloc[0].get("close", 0.0)
                ret1d = univ_match.iloc[0].get("today_return_pct", 0.0)
            elif sym in yfinance_data:
                info = yfinance_data[sym].get("info", {})
                company_name = info.get("longName") or sym
                sector = info.get("sector") or "unknown"
                mcap_cr = info.get("marketCap", 0.0) / 10000000.0 if info.get("marketCap") else 0.0
                close = info.get("previousClose") or 0.0
                
            sig_name = signals_dict.get(sym, {}).get("entry", "Active Signal")
            scans_cnt = scans_dict.get(sym, 0)
            ind_name = industry_dict.get(sym, sector)
            
            confluence_3_rows.append({
                "symbol": sym,
                "company": company_name,
                "industry": ind_name,
                "close": close,
                "return_1d": ret1d,
                "mcap_cr": mcap_cr,
                "signal": sig_name,
                "scans_count": scans_cnt
            })
            
    confluence_3_rows = sorted(confluence_3_rows, key=lambda x: x["scans_count"], reverse=True)
    
    # 2. Collect Emerging Leaders
    emerging_rows = []
    backtest_df = load_backtest_df()
    if not backtest_df.empty:
        emerging_leaders = detect_emerging_leaders(backtest_df)
        
        # Calculate recent counts in last 10 dates
        unique_dates = sorted(backtest_df["parsed_date"].unique())
        recent_dates = unique_dates[-10:] if len(unique_dates) >= 10 else unique_dates
        df_recent = backtest_df[backtest_df["parsed_date"].isin(recent_dates)]
        recent_counts = df_recent["symbol"].value_counts().to_dict()
        
        for sym in emerging_leaders:
            company_name = sym
            sector = "unknown"
            mcap_cr = 0.0
            close = 0.0
            
            univ_match = universe[universe["symbol"] == sym]
            if not univ_match.empty:
                company_name = univ_match.iloc[0]["company"]
                sector = univ_match.iloc[0]["sector"]
                mcap_cr = univ_match.iloc[0].get("marketcap_cr", 0.0)
                close = univ_match.iloc[0].get("close", 0.0)
            elif sym in yfinance_data:
                info = yfinance_data[sym].get("info", {})
                company_name = info.get("longName") or sym
                sector = info.get("sector") or "unknown"
                mcap_cr = info.get("marketCap", 0.0) / 10000000.0 if info.get("marketCap") else 0.0
                close = info.get("previousClose") or 0.0
                
            r_count = recent_counts.get(sym, 0)
            ind_name = industry_dict.get(sym, sector)
            
            emerging_rows.append({
                "symbol": sym,
                "company": company_name,
                "industry": ind_name,
                "close": close,
                "mcap_cr": mcap_cr,
                "persistence": r_count
            })
            
    emerging_rows = sorted(emerging_rows, key=lambda x: x["persistence"], reverse=True)
    
    # 3. Format the Markdown report content
    md = []
    md.append(f"# 📊 Monit High-Conviction Confluences & Emerging Leaders Report")
    md.append(f"Generated on **{today_str}** | Premium Quantitative Watchlist Analysis\n")
    
    md.append(f"## 🏆 1. High-Conviction Confluences (Common Count = 3)")
    md.append(f"These stocks are at the absolute intersection of all three major momentum dimensions:")
    md.append(f"1. **Chartink screener universe** (bullish base)")
    md.append(f"2. **Active scanner signals** (fresh EMA Crossovers, 52W Breakouts, or ATH Momentum)")
    md.append(f"3. **StockScans scan matches** (bullish volume/strength consensus across multiple other watchlists)")
    md.append(f"\nTotal triple-confluence candidates: **{len(confluence_3_rows)}**\n")
    
    if confluence_3_rows:
        md.append("| Symbol | Company Name | Industry | Close (₹) | 1D Ret (%) | Mcap (Cr) | Active Signal | Scans Count | Deep Research Report |")
        md.append("|---|---|---|---|---|---|---|---|---|")
        for r in confluence_3_rows:
            report_status = "_Pending separate pipeline run_"
            if r['symbol'] in existing_reports:
                report_status = f"📝 **Sent on {existing_reports[r['symbol']]['formatted']}**"
                
            md.append(
                f"| `{r['symbol']}` | {r['company']} | {r['industry']} | ₹{r['close']:,.2f} | {r['return_1d']}% | {r['mcap_cr']:,.1f} | **{r['signal']}** | **{r['scans_count']}** | {report_status} |"
            )
    else:
        md.append("_No triple-confluence candidates detected in today's run._")
        
    md.append(f"\n---\n")
    
    md.append(f"## 🚀 2. Emerging Multibagger Leaders (Fresh Momentum Expansion)")
    md.append(f"These stocks show a fresh institutional footprint. They have minimal historical appearances (`<= 2` counts over the prior months) but have erupted recently (`>= 3` appearances in the last 10 days). This highlights **early stage-2 momentum expansion** before they double!")
    md.append(f"\nTotal emerging leaders: **{len(emerging_rows)}**\n")
    
    if emerging_rows:
        md.append("| Rank | Symbol | Company Name | Industry | Close (₹) | Mcap (Cr) | Persistence (Last 10D) |")
        md.append("|---|---|---|---|---|---|---|")
        for idx, r in enumerate(emerging_rows, start=1):
            md.append(
                f"| {idx} | `{r['symbol']}` | {r['company']} | {r['industry']} | ₹{r['close']:,.2f} | {r['mcap_cr']:,.1f} | **{r['persistence']}/10 days** |"
            )
    else:
        md.append("_No emerging leaders detected in today's run._")
        
    CONFLUENCE_EMERGING_REPORT = "\n".join(md)
    
    # Save the report to outputs directory
    try:
        report_path = Path("outputs") / "emerging_and_confluence_report.md"
        report_path.parent.mkdir(parents=True, exist_ok=True)
        with open(report_path, "w") as f:
            f.write(CONFLUENCE_EMERGING_REPORT)
        print(f"✅ Saved automated report to: {report_path}")
    except Exception as e:
        print(f"⚠️ Error saving report to file: {e}")
        
    # Save the confluences list to a JSON file for the separate report pipeline
    try:
        confl_json_path = Path("outputs") / "today_confluences.json"
        with open(confl_json_path, "w") as f:
            json.dump(confluence_3_rows, f, indent=2)
        print(f"✅ Saved today's confluences list to: {confl_json_path}")
    except Exception as e:
        print(f"⚠️ Error saving today's confluences list JSON: {e}")

    # Save the emerging leaders list to a JSON file for the separate report pipeline
    try:
        emerg_json_path = Path("outputs") / "today_emerging.json"
        with open(emerg_json_path, "w") as f:
            json.dump(emerging_rows, f, indent=2)
        print(f"✅ Saved today's emerging leaders list to: {emerg_json_path}")
    except Exception as e:
        print(f"⚠️ Error saving today's emerging leaders list JSON: {e}")
        
    print("✅ Delivery data is fully synchronized in SQLite database.")

        
    # 4. Format the HTML report content for Gmail
    html_confl_rows = ""
    if confluence_3_rows:
        for r in confluence_3_rows:
            report_status = '<span style="color:#777;font-style:italic">Pending separate pipeline</span>'
            if r['symbol'] in existing_reports:
                report_status = f'<b style="color:#2e7d32">📝 Sent on {existing_reports[r["symbol"]]["formatted"]}</b>'
                
            html_confl_rows += f"""
            <tr style="background-color:#fffbee">
              <td style="padding:8px;font-weight:bold;border:1px solid #ddd">🏆 `{r['symbol']}`</td>
              <td style="padding:8px;border:1px solid #ddd">{r['company']}</td>
              <td style="padding:8px;border:1px solid #ddd">{r['industry']}</td>
              <td style="padding:8px;text-align:right;border:1px solid #ddd">₹{r['close']:,.2f}</td>
              <td style="padding:8px;text-align:right;border:1px solid #ddd;color:{'#1a7a1a' if r['return_1d']>=0 else '#cc0000'}">{r['return_1d']}%</td>
              <td style="padding:8px;text-align:right;border:1px solid #ddd">₹{r['mcap_cr']:,.1f}</td>
              <td style="padding:8px;font-weight:bold;border:1px solid #ddd;color:#b25900">{r['signal']}</td>
              <td style="padding:8px;text-align:center;font-weight:bold;border:1px solid #ddd;background-color:#ffe8cc">{r['scans_count']}</td>
              <td style="padding:8px;text-align:center;border:1px solid #ddd;font-size:12px">{report_status}</td>
            </tr>"""
    else:
        html_confl_rows = """<tr><td colspan="9" style="padding:10px;text-align:center;font-style:italic">No triple-confluence candidates detected today.</td></tr>"""
        
    html_emg_rows = ""
    if emerging_rows:
        for idx, r in enumerate(emerging_rows, start=1):
            html_emg_rows += f"""
            <tr style="background-color:#f6f9ff">
              <td style="padding:8px;text-align:center;font-weight:bold;border:1px solid #ddd">{idx}</td>
              <td style="padding:8px;font-weight:bold;border:1px solid #ddd">🚀 `{r['symbol']}`</td>
              <td style="padding:8px;border:1px solid #ddd">{r['company']}</td>
              <td style="padding:8px;border:1px solid #ddd">{r['industry']}</td>
              <td style="padding:8px;text-align:right;border:1px solid #ddd">₹{r['close']:,.2f}</td>
              <td style="padding:8px;text-align:right;border:1px solid #ddd">₹{r['mcap_cr']:,.1f}</td>
              <td style="padding:8px;text-align:center;font-weight:bold;border:1px solid #ddd;background-color:#d0e1fd;color:#004085">{r['persistence']}/10 days</td>
            </tr>"""
    else:
        html_emg_rows = """<tr><td colspan="7" style="padding:10px;text-align:center;font-style:italic">No emerging leaders detected today.</td></tr>"""
        
    CONFLUENCE_EMERGING_HTML = f"""
    <div style="margin-top:25px;border:1px solid #ffe4cc;background-color:#fffaf5;padding:15px;border-radius:6px;font-family:sans-serif">
      <h3 style="color:#c55a11;margin-top:0">🏆 Monit High-Conviction Confluences (Common Count = 3)</h3>
      <p style="font-size:14px;color:#555">These stocks intersect all three core momentum dimensions (Chartink universe + Active scanner signal + StockScans overlap list).</p>
      <table border="1" cellspacing="0" style="border-collapse:collapse;width:100%;font-size:13px;border-color:#ddd">
        <thead style="background-color:#c55a11;color:white">
          <tr>
            <th style="padding:8px">Symbol</th>
            <th style="padding:8px">Company</th>
            <th style="padding:8px">Industry</th>
            <th style="padding:8px">Close</th>
            <th style="padding:8px">1D Ret</th>
            <th style="padding:8px">Mcap Cr</th>
            <th style="padding:8px">Scanner Signal</th>
            <th style="padding:8px">Scans</th>
            <th style="padding:8px">Deep Research Report</th>
          </tr>
        </thead>
        <tbody>
          {html_confl_rows}
        </tbody>
      </table>
    </div>
    
    <div style="margin-top:25px;border:1px solid #d0e1fd;background-color:#fcfdfe;padding:15px;border-radius:6px;font-family:sans-serif">
      <h3 style="color:#2f5597;margin-top:0">🚀 Emerging Multibagger Leaders (Fresh Momentum Expansion)</h3>
      <p style="font-size:14px;color:#555">Stocks with fresh institutional footprint (low historical counts, high recent frequency in the last 10 days). Catching them early before they double!</p>
      <table border="1" cellspacing="0" style="border-collapse:collapse;width:100%;font-size:13px;border-color:#ddd">
        <thead style="background-color:#2f5597;color:white">
          <tr>
            <th style="padding:8px;width:50px">Rank</th>
            <th style="padding:8px">Symbol</th>
            <th style="padding:8px">Company</th>
            <th style="padding:8px">Industry</th>
            <th style="padding:8px">Close</th>
            <th style="padding:8px">Mcap Cr</th>
            <th style="padding:8px">Persistence (10D)</th>
          </tr>
        </thead>
        <tbody>
          {html_emg_rows}
        </tbody>
      </table>
    </div>
    """


def load_scan_matched_symbols() -> list[str]:
    """Read symbols from StockScans live common-stocks data."""
    data = get_stockscans_common_stocks_data()
    companies = data.get("companies", [])
    symbols = []
    for c in companies:
        comp_id = c.get("companyId", "")
        symbol = comp_id.split(":")[1] if ":" in comp_id else comp_id
        if symbol:
            symbols.append(symbol.strip())
    return list(set(symbols))


def load_scan_matched_df(universe: pd.DataFrame, yfinance_data: dict[str, dict]) -> pd.DataFrame:
    """Load StockScans live common-stocks as a normalized DataFrame and enrich sector from Chartink or Yahoo Finance."""
    data = get_stockscans_common_stocks_data()
    companies = data.get("companies", [])
    rows = []
    for c in companies:
        comp_id = c.get("companyId", "")
        symbol = comp_id.split(":")[1] if ":" in comp_id else comp_id
        
        # 1. Try to get sector from Chartink universe first
        sector = "unknown"
        univ_match = universe[universe["symbol"] == symbol]
        if not univ_match.empty and pd.notna(univ_match.iloc[0].get("sector")):
            sector = univ_match.iloc[0]["sector"]
        
        # 2. Try to get sector from Yahoo Finance info second
        if sector == "unknown" or not sector:
            details = yfinance_data.get(symbol, {})
            info = details.get("info", {})
            if info and info.get("sector"):
                sector = info.get("sector")
                
        rows.append({
            "symbol": symbol,
            "company": c.get("Name", ""),
            "sector": sector or "unknown",
            "industry": c.get("Industry", ""),
            "marketcap_bucket": "smallcap" if c.get("Market Capitalization", 0) < 5000 else "midcap" if c.get("Market Capitalization", 0) < 20000 else "largecap",
            "marketcap_cr": c.get("Market Capitalization", 0.0),
            "close": c.get("Close Price", 0.0),
            "today_return_pct": c.get("Returns 1D", 0.0),
            "volume": 0.0,
            "Scans": c.get("Scans", 0)
        })
    return pd.DataFrame(rows)


def build_workbook(
    universe: pd.DataFrame, template_path: Path, output_path: Path
) -> Path:
    # Initialize SQLite database and seed historical CSV if empty, then append today's active signals idempotently
    init_and_seed_database()
    append_today_to_database(universe)

    non_financial_criteria = extract_criteria(template_path, "Non Financial")
    bank_criteria = extract_criteria(template_path, "Banks & NBFC")

    # Load existing inputs from template to preserve scoring!
    non_fin_inputs = extract_existing_company_inputs(template_path, "Non Financial", non_financial_criteria)
    bank_inputs = extract_existing_company_inputs(template_path, "Banks & NBFC", bank_criteria)

    # Fetch yfinance details in parallel for automated filling
    scanner_symbols = load_scanner_symbols()
    scan_matched_symbols = load_scan_matched_symbols()
    symbols = universe["symbol"].dropna().unique().tolist()
    
    # Load backtest database early to identify emerging leaders and fetch their yfinance details
    backtest_df = load_backtest_df()
    emerging_leaders = []
    if not backtest_df.empty:
        try:
            emerging_leaders = detect_emerging_leaders(backtest_df)
        except Exception as e:
            print(f"⚠️ Error detecting emerging leaders early: {e}")
            
    all_symbols = list(set(symbols + scanner_symbols + scan_matched_symbols + emerging_leaders))
    # Fetch benchmarks and StockScans indices first to avoid yfinance rate limiting
    benchmark_data = fetch_benchmark_indices()
    stockscans_indices = fetch_all_stockscans_indices(all_symbols)
    yfinance_data = fetch_all_stocks_details(all_symbols)
    
    # Compute weekly scan match additions and removals
    added_symbols, removed_symbols = track_weekly_scanmatch_changes(scan_matched_symbols)
    
    # Calculate global RS ratings and percentiles
    global_rs = calculate_global_rs_ratings(
        all_symbols=all_symbols,
        universe=universe,
        yfinance_data=yfinance_data,
        benchmark_data=benchmark_data,
        stockscans_indices=stockscans_indices
    )

    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "Read Me"
    write_readme(ws_readme)

    write_universe_sheet(wb.create_sheet("Chartink Universe"), universe)

    banks = universe[universe.apply(is_financial_company, axis=1)].copy()
    non_fin = universe[~universe.apply(is_financial_company, axis=1)].copy()

    monit_non_fin_scores = write_scoring_sheet(
        wb.create_sheet("Monit Non Financial"),
        non_fin,
        non_financial_criteria,
        "Non Financial",
        non_fin_inputs,
        yfinance_data,
        global_rs,
        include_portfolio=False
    )
    monit_bank_scores = write_scoring_sheet(
        wb.create_sheet("Monit Banks NBFC"),
        banks,
        bank_criteria,
        "Banks & NBFC",
        bank_inputs,
        yfinance_data,
        global_rs,
        include_portfolio=False
    )

    # Process and write scanner signals sheets!
    enriched_scanner = []
    for sym in scanner_symbols:
        univ_match = universe[universe["symbol"] == sym]
        if not univ_match.empty:
            row_dict = univ_match.iloc[0].to_dict()
        else:
            details = yfinance_data.get(sym, {})
            info = details.get("info", {})
            row_dict = {
                "symbol": sym,
                "company": info.get("longName", sym),
                "sector": info.get("sector", "unknown"),
                "industry": info.get("industry", "unknown"),
                "marketcap_bucket": "smallcap" if info.get("marketCap", 0) < 50000000000 else "midcap" if info.get("marketCap", 0) < 200000000000 else "largecap",
                "marketcap_cr": round(info.get("marketCap", 0) / 10000000, 2) if info.get("marketCap") else 0.0,
                "close": info.get("previousClose") or info.get("regularMarketPreviousClose") or 0.0,
                "today_return_pct": 0.0,
                "volume": info.get("volume") or 0.0,
            }
        enriched_scanner.append(row_dict)

    scanner_non_fin_scores = {}
    scanner_bank_scores = {}
    if enriched_scanner:
        scanner_df = pd.DataFrame(enriched_scanner)
        scanner_banks = scanner_df[scanner_df.apply(is_financial_company, axis=1)].copy()
        scanner_non_fin = scanner_df[~scanner_df.apply(is_financial_company, axis=1)].copy()

        scanner_non_fin_scores = write_scoring_sheet(
            wb.create_sheet("Scanner Non Financial"),
            scanner_non_fin,
            non_financial_criteria,
            "Non Financial",
            non_fin_inputs,
            yfinance_data,
            global_rs,
            include_portfolio=True
        )
        scanner_bank_scores = write_scoring_sheet(
            wb.create_sheet("Scanner Banks NBFC"),
            scanner_banks,
            bank_criteria,
            "Banks & NBFC",
            bank_inputs,
            yfinance_data,
            global_rs,
            include_portfolio=True
        )

    # Define new custom Criterion for Scan Matches
    scan_match_criterion = Criterion(
        section="TECHNICALS",
        label="Scan Matches",
        question="No. of Scans confluence (>20: 10pts, 10-20: 5pts, <10: 2pts)",
        source_row=999,
        score_formula="=IF(D999>20,10,IF(D999>=10,5,2))",
        validation_formula=None
    )
    scan_match_non_financial_criteria = non_financial_criteria + [scan_match_criterion]
    scan_match_bank_criteria = bank_criteria + [scan_match_criterion]

    # Process and write authenticated Scan Match sheets!
    scan_match_non_fin_scores = {}
    scan_match_bank_scores = {}
    scan_matched_df = load_scan_matched_df(universe, yfinance_data)
    if not scan_matched_df.empty:
        scan_matched_banks = scan_matched_df[scan_matched_df.apply(is_financial_company, axis=1)].copy()
        scan_matched_non_fin = scan_matched_df[~scan_matched_df.apply(is_financial_company, axis=1)].copy()

        scan_match_non_fin_scores = write_scoring_sheet(
            wb.create_sheet("Scan Match Non Financial"),
            scan_matched_non_fin,
            scan_match_non_financial_criteria,
            "Non Financial",
            non_fin_inputs,
            yfinance_data,
            global_rs,
            include_portfolio=False,
            added_symbols=added_symbols,
            removed_symbols=removed_symbols
        )
        scan_match_bank_scores = write_scoring_sheet(
            wb.create_sheet("Scan Match Banks NBFC"),
            scan_matched_banks,
            scan_match_bank_criteria,
            "Banks & NBFC",
            bank_inputs,
            yfinance_data,
            global_rs,
            include_portfolio=False,
            added_symbols=added_symbols,
            removed_symbols=removed_symbols
        )

    # Combine all scores dictionaries!
    all_scores = {}
    all_scores.update(monit_non_fin_scores)
    all_scores.update(monit_bank_scores)
    all_scores.update(scanner_non_fin_scores)
    all_scores.update(scanner_bank_scores)
    all_scores.update(scan_match_non_fin_scores)
    all_scores.update(scan_match_bank_scores)

    # 1. Compile initial scans and industry dictionaries from StockScans 100 list
    scans_dict = {}
    industry_dict = {}
    if not scan_matched_df.empty:
        for _, r in scan_matched_df.iterrows():
            sym = r["symbol"]
            scans_dict[sym] = r.get("Scans") or 0
            industry_dict[sym] = r.get("industry") or "unknown"

    # 2. Compute overlap symbols (Common Count >= 2) to filter out single-match stocks!
    set_universe = set(universe["symbol"].dropna().unique())
    set_scanner = set(scanner_symbols)
    set_scan_match = set(scan_matched_symbols)
    
    overlap_symbols = set()
    union_symbols = set_universe | set_scanner | set_scan_match
    for sym in union_symbols:
        in_univ = sym in set_universe
        in_scan = sym in set_scanner
        in_match = sym in set_scan_match
        if int(in_univ) + int(in_scan) + int(in_match) >= 2:
            overlap_symbols.add(sym)
            
    # Only scrape missing StockScans details for stocks that are part of the overlap!
    missing_stockscans_symbols = list(overlap_symbols - set_scan_match)
    
    stockscans_data = fetch_all_stockscans_details(missing_stockscans_symbols)
    for sym, data in stockscans_data.items():
        if sym not in scans_dict:
            saved_scans = len(data.get("savedScans", []))
            popular_scans = len(data.get("popularScans", []))
            scans_dict[sym] = saved_scans + popular_scans
            
            meta = data.get("metaRatios", {})
            industry_dict[sym] = meta.get("Industry") or "unknown"

    signals_dict = load_scanner_signals_dict()
    backtest_recurrence_dict = load_backtest_recurrence_dict()

    # Create and write Confluence Overlap sheet!
    write_confluence_sheet(
        wb.create_sheet("Confluence Overlap", index=2),
        overlap_symbols,
        universe,
        scanner_symbols,
        scan_matched_symbols,
        yfinance_data,
        all_scores,
        scans_dict,
        signals_dict,
        industry_dict,
        backtest_recurrence_dict
    )

    # Load and build the Multibagger Leaderboard dashboard tab as tab 2!
    if not backtest_df.empty:
        write_leaderboard_sheet(
            wb.create_sheet("Leaderboard", index=1),
            backtest_df,
            universe,
            yfinance_data,
            all_scores,
            global_rs
        )

    write_formula_map(wb.create_sheet("Formula Map"), non_financial_criteria, bank_criteria)

    # Automatically synchronize and update the database rollup analytical profiles
    try:
        rebuild_stock_analytics_table(yfinance_data, scans_dict, industry_dict)
    except Exception as e:
        print(f"⚠️ Error compiling stock analytics: {e}")

    # Automatically generate the Confluence Overlap (Count=3) and Emerging Leaders reports
    try:
        generate_automated_reports(
            universe,
            scanner_symbols,
            scan_matched_symbols,
            yfinance_data,
            scans_dict,
            industry_dict,
            signals_dict
        )
    except Exception as e:
        print(f"⚠️ Error generating automated confluence & emerging report: {e}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def track_weekly_scanmatch_changes(current_symbols: list[str]) -> tuple[set[str], dict[str, str]]:
    """
    Tracks weekly changes (additions/removals) for StockScans 100 stock list.
    Resets every Monday. Added ones are highlighted. Removed ones are listed below the sheet.
    Returns:
        added_symbols: set of symbols currently added (status='added')
        removed_symbols: dict mapping symbol -> removed_date
    """
    import sqlite3
    from datetime import datetime, timedelta
    
    today_dt = datetime.now()
    # Find Monday of this week
    monday_dt = today_dt - timedelta(days=today_dt.weekday())
    week_start_date = monday_dt.strftime("%Y-%m-%d")
    today_str = today_dt.strftime("%Y-%m-%d")
    
    db_path = Path("logs/backtest.db")
    db_path.parent.mkdir(parents=True, exist_ok=True)
    
    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()
    
    # 1. Query records for the current week
    cursor.execute(
        "SELECT symbol, status, added_date, removed_date FROM scanmatch_weekly_tracking WHERE week_start_date = ?",
        (week_start_date,)
    )
    records = cursor.fetchall()
    
    current_symbols_set = {sym.strip().upper() for sym in current_symbols if sym.strip()}
    
    if not records:
        # First run of the week: establish base set
        print(f"🗓️ First run of the week ({week_start_date}). Establishing base set of {len(current_symbols_set)} symbols...")
        for sym in current_symbols_set:
            cursor.execute(
                """INSERT OR REPLACE INTO scanmatch_weekly_tracking 
                   (week_start_date, symbol, status, added_date, removed_date) 
                   VALUES (?, ?, 'base', NULL, NULL)""",
                (week_start_date, sym)
            )
        conn.commit()
        conn.close()
        return set(), {}
        
    # Existent records for this week
    db_status = {row[0]: (row[1], row[2], row[3]) for row in records}
    
    # Process updates:
    for sym in current_symbols_set:
        if sym in db_status:
            status, added_date, removed_date = db_status[sym]
            if status == "removed":
                # Restore to base or added depending on if added_date was set
                orig_status = "added" if added_date else "base"
                cursor.execute(
                    """UPDATE scanmatch_weekly_tracking 
                       SET status = ?, removed_date = NULL 
                       WHERE week_start_date = ? AND symbol = ?""",
                    (orig_status, week_start_date, sym)
                )
        else:
            # New addition
            cursor.execute(
                """INSERT INTO scanmatch_weekly_tracking 
                   (week_start_date, symbol, status, added_date, removed_date) 
                   VALUES (?, ?, 'added', ?, NULL)""",
                (week_start_date, sym, today_str)
            )
            
    # Process removals
    for sym, (status, added_date, removed_date) in db_status.items():
        if status in ("base", "added") and sym not in current_symbols_set:
            cursor.execute(
                """UPDATE scanmatch_weekly_tracking 
                   SET status = 'removed', removed_date = ? 
                   WHERE week_start_date = ? AND symbol = ?""",
                (today_str, week_start_date, sym)
            )
            
    conn.commit()
    
    # Fetch final updated set of records
    cursor.execute(
        "SELECT symbol, status, added_date, removed_date FROM scanmatch_weekly_tracking WHERE week_start_date = ?",
        (week_start_date,)
    )
    records = cursor.fetchall()
    conn.close()
    
    added_symbols = set()
    removed_symbols = {}
    
    for symbol, status, added_date, removed_date in records:
        if status == "added":
            added_symbols.add(symbol)
        elif status == "removed":
            removed_symbols[symbol] = removed_date
            
    return added_symbols, removed_symbols



def init_and_seed_database() -> None:
    """Initialize local SQLite database logs/backtest.db and seed historical 9-month CSV if empty."""
    db_path = Path("logs/backtest.db")
    db_path.parent.mkdir(parents=True, exist_ok=True)
    
    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS screener_history (
            date TEXT,
            symbol TEXT,
            marketcapname TEXT,
            sector TEXT,
            PRIMARY KEY (date, symbol)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS stock_analytics (
            symbol TEXT PRIMARY KEY,
            company TEXT,
            sector TEXT,
            industry TEXT,
            marketcap_cr REAL,
            scan_match_count INTEGER,
            first_appeared_date TEXT,
            first_appeared_price REAL,
            current_price REAL,
            price_diff_pct REAL,
            total_appearances INTEGER,
            current_streak INTEGER,
            is_emerging INTEGER,
            appearances_15d INTEGER,
            appearances_30d INTEGER,
            appearances_90d INTEGER
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS scanmatch_weekly_tracking (
            week_start_date TEXT,
            symbol TEXT,
            status TEXT,
            added_date TEXT,
            removed_date TEXT,
            PRIMARY KEY (week_start_date, symbol)
        )
    """)
    conn.commit()
    
    cursor.execute("SELECT COUNT(*) FROM screener_history")
    count = cursor.fetchone()[0]
    
    if count == 0:
        csv_path = Path("Backtest Monit momentum (2).csv")
        if csv_path.exists():
            print(f"🌱 Seeding database with historical 9-month data from '{csv_path}'...")
            try:
                df = pd.read_csv(csv_path)
                df = df.dropna(subset=["date", "symbol"])
                df["date"] = df["date"].astype(str).str.strip()
                df["symbol"] = df["symbol"].astype(str).str.strip().str.upper()
                df["marketcapname"] = df["marketcapname"].astype(str).str.strip()
                df["sector"] = df["sector"].astype(str).str.strip()
                
                # Deduplicate to prevent constraint failures
                df = df.drop_duplicates(subset=["date", "symbol"])
                
                records = [
                    (row.date, row.symbol, row.marketcapname, row.sector)
                    for row in df.itertuples(index=False)
                ]
                
                cursor.executemany(
                    "INSERT OR IGNORE INTO screener_history (date, symbol, marketcapname, sector) VALUES (?, ?, ?, ?)",
                    records
                )
                conn.commit()
                print(f"✅ Seeding complete! Inserted {len(records)} records into SQLite database.")
            except Exception as e:
                print(f"⚠️ Error seeding SQLite database: {e}")
        else:
            print("⚠️ SQLite database is empty and no baseline CSV 'Backtest Monit momentum (2).csv' was found to seed from.")
            
    conn.close()
    
    # Initialize and sync delivery history database
    try:
        from generate_equity_reports import sync_delivery_history
        print("🔄 Synchronizing local delivery history database from bulk bhavcopy...")
        sync_delivery_history(45)
    except Exception as e:
        print(f"⚠️ Error initializing/syncing delivery history database: {e}")



def append_today_to_database(universe_df: pd.DataFrame) -> None:
    """Idempotently append today's active screener universe signals into SQLite database."""
    if universe_df.empty:
        print("⚠️ Chartink universe is empty today. Skipping database append.")
        return
        
    db_path = Path("logs/backtest.db")
    init_and_seed_database()
    
    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()
    
    today_str = date.today().strftime("%d-%m-%Y")
    
    # Idempotent delete first to prevent duplicate entries for today
    cursor.execute("DELETE FROM screener_history WHERE date = ?", (today_str,))
    deleted = cursor.rowcount
    if deleted > 0:
        print(f"🗑️ Cleaned {deleted} duplicate records for today ({today_str}) in SQLite database.")
        
    # Prepare records
    records = []
    for _, row in universe_df.iterrows():
        sym = str(row.get("symbol", "")).strip().upper()
        if not sym:
            continue
        mcap = str(row.get("marketcap_bucket", "Smallcap")).strip()
        # Clean capitalization of marketcap name
        if mcap.lower() == "largecap":
            mcap = "Largecap"
        elif mcap.lower() == "midcap":
            mcap = "Midcap"
        elif mcap.lower() == "smallcap":
            mcap = "Smallcap"
        else:
            mcap = mcap.capitalize()
            
        sector = str(row.get("sector", "unknown")).strip()
        records.append((today_str, sym, mcap, sector))
        
    if records:
        cursor.executemany(
            "INSERT INTO screener_history (date, symbol, marketcapname, sector) VALUES (?, ?, ?, ?)",
            records
        )
        conn.commit()
        print(f"📝 Successfully appended {len(records)} active signals for today ({today_str}) into SQLite database.")
        
    conn.close()


def load_backtest_df() -> pd.DataFrame:
    """Load backtest data from SQLite database, falling back to CSV if missing."""
    db_path = Path("logs/backtest.db")
    if not db_path.exists():
        csv_path = Path("Backtest Monit momentum (2).csv")
        if csv_path.exists():
            init_and_seed_database()
        else:
            return pd.DataFrame()
            
    try:
        conn = sqlite3.connect(str(db_path))
        df = pd.read_sql_query("SELECT date, symbol, marketcapname, sector FROM screener_history", conn)
        conn.close()
        df["parsed_date"] = pd.to_datetime(df["date"], format="%d-%m-%Y", errors="coerce")
        df = df.dropna(subset=["parsed_date"])
        return df.sort_values("parsed_date").reset_index(drop=True)
    except Exception as e:
        print(f"⚠️ Error reading SQLite database: {e}. Falling back to CSV.")
        csv_path = Path("Backtest Monit momentum (2).csv")
        if not csv_path.exists():
            return pd.DataFrame()
        try:
            df = pd.read_csv(csv_path)
            df["parsed_date"] = pd.to_datetime(df["date"], format="%d-%m-%Y", errors="coerce")
            df = df.dropna(subset=["parsed_date"])
            return df.sort_values("parsed_date").reset_index(drop=True)
        except Exception:
            return pd.DataFrame()


def fetch_historical_prices(symbols: list[str]) -> dict[str, pd.Series]:
    """Fetch up to 2y daily close prices for multiple symbols in parallel."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    results = {}
    
    def worker(sym: str):
        # gentle sleep to rate limit
        time.sleep(0.01)
        for suffix in [".NS", ".BO"]:
            ticker = sym + suffix
            try:
                df = yf.download(ticker, period="2y", interval="1d", progress=False, auto_adjust=True, threads=False)
                if df is not None and not df.empty:
                    # Normalize index
                    if isinstance(df.columns, pd.MultiIndex):
                        if ticker in df.columns.get_level_values(-1):
                            df = df.xs(ticker, axis=1, level=-1)
                        else:
                            df.columns = [c[0] for c in df.columns]
                    return sym, df["Close"].squeeze()
            except Exception:
                pass
        return sym, None
        
    print(f"📥 Downloading historical daily close prices for {len(symbols)} unique symbols in parallel (30 threads)...")
    with ThreadPoolExecutor(max_workers=30) as ex:
        futures = {ex.submit(worker, s): s for s in symbols}
        for i, f in enumerate(as_completed(futures), 1):
            sym = futures[f]
            try:
                symbol, close_series = f.result()
                if close_series is not None and not close_series.empty:
                    results[symbol] = close_series
                if i % 50 == 0 or i == len(symbols):
                     print(f"  [{i}/{len(symbols)}] Fetched price history for {sym}")
            except Exception:
                pass
                
    return results


def to_float_scalar(val) -> float:
    if val is None:
        return 0.0
    if hasattr(val, "values"):
        if len(val) > 0:
            val = val.values[0]
        else:
            return 0.0
    try:
        return float(val)
    except Exception:
        return 0.0


def rebuild_stock_analytics_table(yfinance_data: dict, scans_dict: dict, industry_dict: dict) -> None:
    """Compile analytical insights from raw history ledger and yfinance/StockScans, then update SQLite."""
    print("🔄 Synchronizing and updating 'stock_analytics' rollup table...")
    db_path = Path("logs/backtest.db")
    if not db_path.exists():
        print("⚠️ Database does not exist. Cannot build analytics.")
        return
        
    df_history = load_backtest_df()
    if df_history.empty:
        print("⚠️ No history records found in SQLite database. Skipping analytics rollup.")
        return
        
    conn = sqlite3.connect(str(db_path))
    
    # 1. Load existing cache from database to preserve StockScans industry & match counts
    existing_cache = {}
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT symbol, industry, scan_match_count, company, sector, marketcap_cr FROM stock_analytics")
        for row in cursor.fetchall():
            existing_cache[row[0]] = {
                "industry": row[1],
                "scan_match_count": row[2],
                "company": row[3],
                "sector": row[4],
                "marketcap_cr": row[5]
            }
    except Exception:
        pass
        
    # Find all unique symbols
    unique_symbols = df_history["symbol"].dropna().unique().tolist()
    
    # Calculate earliest appearance per symbol
    df_first = df_history.drop_duplicates(subset=["symbol"], keep="first")
    first_dates = {}
    for _, row in df_first.iterrows():
        first_dates[row["symbol"]] = row["parsed_date"]
        
    # Calculate latest appearance per symbol (for current streak, latest sector, etc.)
    df_last = df_history.drop_duplicates(subset=["symbol"], keep="last")
    sectors = {}
    for _, row in df_last.iterrows():
        sectors[row["symbol"]] = row["sector"]
        
    # Calculate total appearances
    df_counts = df_history.groupby("symbol")["date"].nunique().to_dict()
    
    # Calculate rolling counts (15 days, 30 days, 90 days)
    unique_dates = sorted(df_history["parsed_date"].unique())
    last_15_dates = unique_dates[-15:] if len(unique_dates) >= 15 else unique_dates
    last_30_dates = unique_dates[-30:] if len(unique_dates) >= 30 else unique_dates
    last_90_dates = unique_dates[-90:] if len(unique_dates) >= 90 else unique_dates
    
    df_15d = df_history[df_history["parsed_date"].isin(last_15_dates)].groupby("symbol")["date"].nunique().to_dict()
    df_30d = df_history[df_history["parsed_date"].isin(last_30_dates)].groupby("symbol")["date"].nunique().to_dict()
    df_90d = df_history[df_history["parsed_date"].isin(last_90_dates)].groupby("symbol")["date"].nunique().to_dict()
    
    # Calculate streaks and emerging leaders
    streaks = calculate_consecutive_streaks(df_history)
    emerging_leaders = set(detect_emerging_leaders(df_history))
    
    # 2. Determine which symbols need price history download
    historical_closes = fetch_historical_prices(unique_symbols)
    
    # 3. Compile final rows for insertion
    insert_records = []
    for sym in unique_symbols:
        first_date = first_dates.get(sym)
        first_date_str = first_date.strftime("%d-%m-%Y") if first_date else "unknown"
        
        # Determine company name, sector, exact market cap, industry, and scan match count
        cache = existing_cache.get(sym) or {}
        
        # Company & Sector
        company_name = cache.get("company")
        sector = sectors.get(sym) or cache.get("sector") or "unknown"
        
        # Check today's yfinance data
        yf_details = yfinance_data.get(sym) or {}
        yf_info = yf_details.get("info") or {}
        
        if yf_info:
            if not company_name:
                company_name = yf_info.get("longName") or yf_info.get("shortName") or sym
            marketcap_cr = round(yf_info.get("marketCap", 0) / 10000000, 2) if yf_info.get("marketCap") else cache.get("marketcap_cr") or 0.0
        else:
            marketcap_cr = cache.get("marketcap_cr") or 0.0
            
        if not company_name:
            company_name = sym # Fallback
            
        # Industry & Scan Match Count (strictly from StockScans)
        # If in today's scraped industry_dict, use it. Otherwise, use existing cached value.
        industry = industry_dict.get(sym) or cache.get("industry") or "unknown"
        scan_match_count = scans_dict.get(sym) or cache.get("scan_match_count") or 0
        
        # Calculate Prices & Returns
        first_price = 0.0
        current_price = 0.0
        price_diff_pct = 0.0
        
        closes = historical_closes.get(sym)
        if closes is not None and not closes.empty:
            if isinstance(closes, pd.DataFrame):
                closes = closes.iloc[:, 0]
                
            current_price = to_float_scalar(closes.iloc[-1])
            if first_date:
                matching_dates = closes.index[closes.index >= first_date]
                if not matching_dates.empty:
                    first_price = to_float_scalar(closes.loc[matching_dates[0]])
                else:
                    first_price = to_float_scalar(closes.iloc[-1])
                    
            if first_price > 0:
                price_diff_pct = round(((current_price - first_price) / first_price) * 100, 2)
                
        # Rollup Metrics
        total_apps = df_counts.get(sym, 0)
        streak = streaks.get(sym, 0)
        is_emerging = 1 if sym in emerging_leaders else 0
        
        apps_15d = df_15d.get(sym, 0)
        apps_30d = df_30d.get(sym, 0)
        apps_90d = df_90d.get(sym, 0)
        
        insert_records.append((
            sym,
            company_name,
            sector,
            industry,
            marketcap_cr,
            scan_match_count,
            first_date_str,
            round(first_price, 2),
            round(current_price, 2),
            price_diff_pct,
            total_apps,
            streak,
            is_emerging,
            apps_15d,
            apps_30d,
            apps_90d
        ))
        
    # Write to database
    if insert_records:
        cursor = conn.cursor()
        cursor.executemany("""
            INSERT OR REPLACE INTO stock_analytics (
                symbol, company, sector, industry, marketcap_cr, scan_match_count,
                first_appeared_date, first_appeared_price, current_price, price_diff_pct,
                total_appearances, current_streak, is_emerging,
                appearances_15d, appearances_30d, appearances_90d
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, insert_records)
        conn.commit()
        print(f"✅ Successfully updated and synced {len(insert_records)} analytical profiles in 'stock_analytics' table.")
        
    conn.close()


def calculate_consecutive_streaks(df_backtest: pd.DataFrame) -> dict[str, int]:
    """Calculate the consecutive trading day streak for each symbol ending at the latest backtest date."""
    if df_backtest.empty:
        return {}
    unique_dates = sorted(df_backtest["parsed_date"].unique())
    if not unique_dates:
        return {}
    latest_date = unique_dates[-1]
    symbol_dates = {}
    for _, r in df_backtest.iterrows():
        sym = r["symbol"]
        symbol_dates.setdefault(sym, set()).add(r["parsed_date"])
    streaks = {}
    for sym, dates in symbol_dates.items():
        streak = 0
        for d in reversed(unique_dates):
            if d in dates:
                streak += 1
            else:
                break
        streaks[sym] = streak
    return streaks


def calculate_sector_rotation_velocity(df_backtest: pd.DataFrame) -> pd.DataFrame:
    """Calculate WoW Sector Momentum Velocity by comparing the last 5 trading days vs previous 5 trading days."""
    if df_backtest.empty:
        return pd.DataFrame()
    unique_dates = sorted(df_backtest["parsed_date"].unique())
    if len(unique_dates) < 10:
        return pd.DataFrame()
    week1_dates = unique_dates[-5:]
    week2_dates = unique_dates[-10:-5]
    df_w1 = df_backtest[df_backtest["parsed_date"].isin(week1_dates)]
    df_w2 = df_backtest[df_backtest["parsed_date"].isin(week2_dates)]
    w1_counts = df_w1["sector"].value_counts().to_dict()
    w2_counts = df_w2["sector"].value_counts().to_dict()
    all_sectors = set(w1_counts.keys()) | set(w2_counts.keys())
    rows = []
    for sector in all_sectors:
        c1 = w1_counts.get(sector, 0)
        c2 = w2_counts.get(sector, 0)
        if c2 == 0:
            velocity = 100.0 if c1 > 0 else 0.0
        else:
            velocity = round(((c1 - c2) / c2) * 100.0, 1)
        rows.append({
            "sector": sector,
            "current_week_count": c1,
            "previous_week_count": c2,
            "velocity_pct": velocity
        })
    return pd.DataFrame(rows).sort_values("velocity_pct", ascending=False).reset_index(drop=True)


def detect_emerging_leaders(df_backtest: pd.DataFrame) -> list[str]:
    """Identify symbols showing fresh momentum expansion (low historical counts, high recent counts)."""
    if df_backtest.empty:
        return []
    unique_dates = sorted(df_backtest["parsed_date"].unique())
    if len(unique_dates) < 15:
        return []
    recent_dates = unique_dates[-10:]
    prior_dates = unique_dates[-60:-10] if len(unique_dates) >= 60 else unique_dates[:-10]
    df_recent = df_backtest[df_backtest["parsed_date"].isin(recent_dates)]
    df_prior = df_backtest[df_backtest["parsed_date"].isin(prior_dates)]
    recent_counts = df_recent["symbol"].value_counts().to_dict()
    prior_counts = df_prior["symbol"].value_counts().to_dict()
    leaders = []
    for sym, r_count in recent_counts.items():
        p_count = prior_counts.get(sym, 0)
        if p_count <= 2 and r_count >= 3:
            leaders.append(sym)
    return leaders


def write_leaderboard_sheet(
    ws,
    df_backtest: pd.DataFrame,
    universe: pd.DataFrame,
    yfinance_data: dict[str, dict],
    all_scores: dict[str, float],
    global_rs: dict[str, dict]
) -> None:
    ws.sheet_view.showGridLines = False
    
    # Title Block
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:N2")
    title_cell = ws["A1"]
    title_cell.value = "🌟 MULTIBAGGER MOMENTUM & SECTOR ROTATION LEADERBOARD 🌟"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Styling helpers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    sub_header_fill = PatternFill("solid", fgColor="2F5597")
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3")
    )
    
    # ─── PRE-COMPUTE METRICS ──────────────────────────────────────────────
    streaks = calculate_consecutive_streaks(df_backtest)
    
    unique_dates = sorted(df_backtest["parsed_date"].unique())
    last_30_dates = unique_dates[-30:] if len(unique_dates) >= 30 else unique_dates
    df_30 = df_backtest[df_backtest["parsed_date"].isin(last_30_dates)]
    counts_30d = df_30.groupby("symbol")["date"].nunique().to_dict()
    
    persistence_scores = {}
    for sym in counts_30d.keys():
        c_30 = counts_30d.get(sym, 0)
        streak = streaks.get(sym, 0)
        score = (c_30 * 3) + (streak * 6)
        persistence_scores[sym] = score
        
    sorted_persistent_syms = sorted(persistence_scores.items(), key=lambda x: x[1], reverse=True)
    df_velocity = calculate_sector_rotation_velocity(df_backtest)
    emerging_leaders = detect_emerging_leaders(df_backtest)
    
    # ─── TABLE 1: TOP PERSISTENT LEADERS (LEFT SIDE: COL A-H) ─────────────
    ws.cell(4, 1, "TOP PERSISTENT MOMENTUM LEADERS (30D)").font = Font(name="Calibri", size=12, bold=True, color="1F4E78")
    t1_headers = ["Rank", "Score", "Symbol", "Company Name", "Sector", "Streak (Days)", "Market Cap Cr", "Classification"]
    for col_idx, h in enumerate(t1_headers, start=1):
        cell = ws.cell(5, col_idx, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    t1_row = 6
    for idx, (sym, p_score) in enumerate(sorted_persistent_syms[:20], 1):
        comp_name = sym
        sector = "unknown"
        mcap_cr = 0.0
        
        univ_match = universe[universe["symbol"] == sym]
        if not univ_match.empty:
            comp_name = univ_match.iloc[0]["company"]
            sector = univ_match.iloc[0]["sector"]
            mcap_cr = univ_match.iloc[0]["marketcap_cr"]
        else:
            details = yfinance_data.get(sym, {})
            info = details.get("info", {})
            if info:
                comp_name = info.get("longName", sym)
                sector = info.get("sector", "unknown")
                mcap_cr = round(info.get("marketCap", 0) / 10000000, 2) if info.get("marketCap") else 0.0
                
        streak = streaks.get(sym, 0)
        classification = "Potential Leader" if streak > 10 else "Institutional Trend" if streak > 5 else "Momentum"
        
        ws.cell(t1_row, 1, idx).border = thin_border
        ws.cell(t1_row, 1).alignment = Alignment(horizontal="center")
        ws.cell(t1_row, 2, p_score).font = Font(bold=True)
        ws.cell(t1_row, 2).border = thin_border
        ws.cell(t1_row, 2).alignment = Alignment(horizontal="center")
        ws.cell(t1_row, 3, sym).font = Font(bold=True)
        ws.cell(t1_row, 3).border = thin_border
        ws.cell(t1_row, 3).alignment = Alignment(horizontal="center")
        ws.cell(t1_row, 4, comp_name).border = thin_border
        ws.cell(t1_row, 5, sector).border = thin_border
        ws.cell(t1_row, 6, streak).border = thin_border
        ws.cell(t1_row, 6).alignment = Alignment(horizontal="center")
        ws.cell(t1_row, 7, mcap_cr).border = thin_border
        ws.cell(t1_row, 7).number_format = "#,##0.0"
        ws.cell(t1_row, 7).alignment = Alignment(horizontal="right")
        ws.cell(t1_row, 8, classification).border = thin_border
        t1_row += 1
        
    # ─── TABLE 3: SECTOR MOMENTUM HEATMAP (RIGHT SIDE: COL J-M) ───────────
    ws.cell(4, 10, "SECTOR MOMENTUM VELOCITY HEATMAP (WoW)").font = Font(name="Calibri", size=12, bold=True, color="1F4E78")
    t3_headers = ["Sector Name", "Current Wk Count", "Previous Wk Count", "Velocity WoW %"]
    for col_idx, h in enumerate(t3_headers, start=10):
        cell = ws.cell(5, col_idx, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    t3_row = 6
    if not df_velocity.empty:
        for _, r in df_velocity.head(15).iterrows():
            ws.cell(t3_row, 10, r["sector"]).border = thin_border
            ws.cell(t3_row, 11, r["current_week_count"]).border = thin_border
            ws.cell(t3_row, 11).alignment = Alignment(horizontal="center")
            ws.cell(t3_row, 12, r["previous_week_count"]).border = thin_border
            ws.cell(t3_row, 12).alignment = Alignment(horizontal="center")
            
            vel_cell = ws.cell(t3_row, 13, f"{r['velocity_pct']}%" if r["velocity_pct"] >= 0 else f"{r['velocity_pct']}%")
            vel_cell.border = thin_border
            vel_cell.alignment = Alignment(horizontal="center")
            vel_cell.font = Font(bold=True, color="006100" if r["velocity_pct"] >= 0 else "9C0006")
            vel_cell.fill = PatternFill("solid", fgColor="C6EFCE" if r["velocity_pct"] >= 0 else "FFC7CE")
            t3_row += 1
            
    # ─── TABLE 2: EMERGING LEADERS (LOWER LEFT: COL A-F, ROW 28 ONWARDS) ────
    start_r2 = max(t1_row + 2, 28)
    ws.cell(start_r2, 1, "🌟 EMERGING LEADERS (MOMENTUM EXPANSION)").font = Font(name="Calibri", size=12, bold=True, color="C55A11")
    t2_headers = ["Symbol", "Company Name", "Sector", "Recent Matches (10D)", "Historical Matches", "Market Cap Cr"]
    for col_idx, h in enumerate(t2_headers, start=1):
        cell = ws.cell(start_r2 + 1, col_idx, h)
        cell.font = header_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    t2_row = start_r2 + 2
    if emerging_leaders:
        recent_dates_10 = unique_dates[-10:]
        prior_dates_60 = unique_dates[-60:-10] if len(unique_dates) >= 60 else unique_dates[:-10]
        df_rec = df_backtest[df_backtest["parsed_date"].isin(recent_dates_10)]
        df_pri = df_backtest[df_backtest["parsed_date"].isin(prior_dates_60)]
        rec_counts = df_rec["symbol"].value_counts().to_dict()
        pri_counts = df_pri["symbol"].value_counts().to_dict()
        
        for sym in emerging_leaders[:12]:
            comp_name = sym
            sector = "unknown"
            mcap_cr = 0.0
            
            univ_match = universe[universe["symbol"] == sym]
            if not univ_match.empty:
                comp_name = univ_match.iloc[0]["company"]
                sector = univ_match.iloc[0]["sector"]
                mcap_cr = univ_match.iloc[0]["marketcap_cr"]
            else:
                details = yfinance_data.get(sym, {})
                info = details.get("info", {})
                if info:
                    comp_name = info.get("longName", sym)
                    sector = info.get("sector", "unknown")
                    mcap_cr = round(info.get("marketCap", 0) / 10000000, 2) if info.get("marketCap") else 0.0
                    
            ws.cell(t2_row, 1, sym).font = Font(bold=True)
            ws.cell(t2_row, 1).border = thin_border
            ws.cell(t2_row, 1).alignment = Alignment(horizontal="center")
            ws.cell(t2_row, 2, comp_name).border = thin_border
            ws.cell(t2_row, 3, sector).border = thin_border
            ws.cell(t2_row, 4, rec_counts.get(sym, 0)).border = thin_border
            ws.cell(t2_row, 4).alignment = Alignment(horizontal="center")
            ws.cell(t2_row, 5, pri_counts.get(sym, 0)).border = thin_border
            ws.cell(t2_row, 5).alignment = Alignment(horizontal="center")
            ws.cell(t2_row, 6, mcap_cr).border = thin_border
            ws.cell(t2_row, 6).number_format = "#,##0.0"
            ws.cell(t2_row, 6).alignment = Alignment(horizontal="right")
            t2_row += 1
    else:
        ws.cell(t2_row, 1, "No emerging leaders detected (requires fresh accumulation footprint).").font = Font(italic=True)
        
    # ─── TABLE 4: SATELLITE SHORTLIST (LOWER RIGHT: COL J-N, ROW 28 ONWARDS) ──
    start_r4 = max(t3_row + 2, 28)
    ws.cell(start_r4, 10, "💎 SATELLITE HOLDINGS SHORTLIST (MCAP < 25K CR)").font = Font(name="Calibri", size=12, bold=True, color="7030A0")
    t4_headers = ["Symbol", "Sector", "Monit Score", "Market Cap Cr", "Appearances (30D)"]
    for col_idx, h in enumerate(t4_headers, start=10):
        cell = ws.cell(start_r4 + 1, col_idx, h)
        cell.font = header_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    t4_row = start_r4 + 2
    sweet_spot_candidates = []
    for sym, p_score in sorted_persistent_syms:
        mcap_cr = 0.0
        sector = "unknown"
        univ_match = universe[universe["symbol"] == sym]
        if not univ_match.empty:
            mcap_cr = univ_match.iloc[0]["marketcap_cr"]
            sector = univ_match.iloc[0]["sector"]
        else:
            details = yfinance_data.get(sym, {})
            info = details.get("info", {})
            if info:
                mcap_cr = round(info.get("marketCap", 0) / 10000000, 2) if info.get("marketCap") else 0.0
                sector = info.get("sector", "unknown")
                
        if 500.0 <= mcap_cr <= 25000.0:
            sweet_spot_candidates.append({
                "symbol": sym,
                "sector": sector,
                "score": all_scores.get(sym, 0.0),
                "mcap_cr": mcap_cr,
                "counts_30d": counts_30d.get(sym, 0)
            })
            
    sweet_spot_candidates = sorted(sweet_spot_candidates, key=lambda x: x["score"], reverse=True)
    
    for item in sweet_spot_candidates[:12]:
        ws.cell(t4_row, 10, item["symbol"]).font = Font(bold=True)
        ws.cell(t4_row, 10).border = thin_border
        ws.cell(t4_row, 10).alignment = Alignment(horizontal="center")
        ws.cell(t4_row, 11, item["sector"]).border = thin_border
        ws.cell(t4_row, 12, item["score"]).border = thin_border
        ws.cell(t4_row, 12).number_format = "#,##0.0"
        ws.cell(t4_row, 12).alignment = Alignment(horizontal="right")
        ws.cell(t4_row, 13, item["mcap_cr"]).border = thin_border
        ws.cell(t4_row, 13).number_format = "#,##0.0"
        ws.cell(t4_row, 13).alignment = Alignment(horizontal="right")
        ws.cell(t4_row, 14, item["counts_30d"]).border = thin_border
        ws.cell(t4_row, 14).alignment = Alignment(horizontal="center")
        t4_row += 1

    # ─── TABLE 5: HIGH MOMENTUM LEADERS (RS RATING >= 90) (LOWER SIDE: COL A-H, ROW 44+) ────
    start_r5 = max(max(t2_row, t4_row) + 3, 44)
    ws.cell(start_r5, 1, "🚀 HIGH MOMENTUM LEADERS (RS RATING >= 90)").font = Font(name="Calibri", size=12, bold=True, color="C00000")
    t5_headers = ["Rank", "RS Rating", "O'Neil Score", "Monit Score", "Symbol", "Company Name", "Sector", "Market Cap Cr"]
    for col_idx, h in enumerate(t5_headers, start=1):
        cell = ws.cell(start_r5 + 1, col_idx, h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="C00000") # Crimson header for high momentum!
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    rs_leaders = []
    for sym, metrics in global_rs.items():
        rating = metrics.get("rs_rating")
        if rating and isinstance(rating, int) and rating >= 90:
            rs_leaders.append({
                "symbol": sym,
                "rs_rating": rating,
                "weighted_score": metrics.get("weighted_score"),
                "monit_score": all_scores.get(sym, 0.0)
            })
            
    # Sort rs_leaders by rs_rating descending, then by weighted_score descending
    rs_leaders = sorted(rs_leaders, key=lambda x: (x["rs_rating"], x["weighted_score"] or 0.0), reverse=True)
    
    t5_row = start_r5 + 2
    for idx, item in enumerate(rs_leaders, start=1):
        sym = item["symbol"]
        comp_name = sym
        sector = "unknown"
        mcap_cr = 0.0
        
        univ_match = universe[universe["symbol"] == sym]
        if not univ_match.empty:
            comp_name = univ_match.iloc[0]["company"]
            sector = univ_match.iloc[0]["sector"]
            mcap_cr = univ_match.iloc[0]["marketcap_cr"]
        else:
            details = yfinance_data.get(sym, {})
            info = details.get("info", {})
            if info:
                comp_name = info.get("longName", sym)
                sector = info.get("sector", "unknown")
                mcap_cr = round(info.get("marketCap", 0) / 10000000, 2) if info.get("marketCap") else 0.0
                
        ws.cell(t5_row, 1, idx).border = thin_border
        ws.cell(t5_row, 1).alignment = Alignment(horizontal="center")
        
        # Color coding cell based on rating
        rating = item["rs_rating"]
        rating_cell = ws.cell(t5_row, 2, rating)
        rating_cell.border = thin_border
        rating_cell.alignment = Alignment(horizontal="center")
        if rating >= 95:
            rating_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            rating_cell.font = Font(color="006100", bold=True)
        else:
            rating_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            rating_cell.font = Font(color="375623", bold=False)
            
        ws.cell(t5_row, 3, item["weighted_score"]).border = thin_border
        ws.cell(t5_row, 3).alignment = Alignment(horizontal="right")
        ws.cell(t5_row, 3).number_format = "#,##0.00"
        
        ws.cell(t5_row, 4, item["monit_score"]).border = thin_border
        ws.cell(t5_row, 4).alignment = Alignment(horizontal="right")
        ws.cell(t5_row, 4).number_format = "#,##0.00"
        
        ws.cell(t5_row, 5, sym).font = Font(bold=True)
        ws.cell(t5_row, 5).border = thin_border
        ws.cell(t5_row, 5).alignment = Alignment(horizontal="center")
        
        ws.cell(t5_row, 6, comp_name).border = thin_border
        ws.cell(t5_row, 7, sector).border = thin_border
        
        ws.cell(t5_row, 8, mcap_cr).border = thin_border
        ws.cell(t5_row, 8).number_format = "#,##0.0"
        ws.cell(t5_row, 8).alignment = Alignment(horizontal="right")
        t5_row += 1
        
    for col in range(1, 15):
        letter = get_column_letter(col)
        max_len = 0
        for r in range(4, max(t2_row, t4_row, t5_row) + 1):
            val = ws.cell(r, col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = max(max_len + 3, 11)


def write_readme(ws) -> None:
    # Set sheet view grids visible
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws.cell(1, 1, "🏆 MONIT MULTIBAGGER RESEARCH ENGINE & SCANNER").font = Font(name="Segoe UI", size=18, bold=True, color="1B365D")
    ws.cell(2, 1, f"Daily Analytics Report | Generated: {date.today().strftime('%A, %d %B %Y')}").font = Font(name="Segoe UI", size=11, italic=True, color="555555")
    
    # Divider line
    for col in range(1, 4):
        ws.cell(3, col).border = Border(bottom=Side(style="medium", color="1B365D"))
        
    # Quick Navigation / How to use section
    ws.cell(5, 1, "📖 QUICK REFERENCE RUNBOOK & SYSTEM FLOW").font = Font(name="Segoe UI", size=12, bold=True, color="1B365D")
    
    how_to_rows = [
        ("Step 1: Reference this Read Me", "Understand terms, formulas, and underlying investment theses directly in this glossary dashboard."),
        ("Step 2: Check Leaderboard Tab", "Locate long-term compounders with high Persistence Scores and streaks (institutional accumulation footprint)."),
        ("Step 3: Analyze Confluence Overlap", "Review the ultimate high-conviction candidates intersecting our scanner signals and StockScans overlaps."),
        ("Step 4: Scoring Monit Tabs", "Open 'Monit Non Financial' or 'Monit Banks NBFC' to fill out your qualitative checks (Tailwinds, Promoters, Triggers)."),
        ("Step 5: Track Scan Matches", "Confirm consensus using 'Scan Match' tabs which score companies by total appearances across allStockScans."),
    ]
    
    r = 6
    for title, desc in how_to_rows:
        ws.cell(r, 1, title).font = Font(name="Segoe UI", size=10, bold=True, color="333333")
        ws.cell(r, 2, desc).font = Font(name="Segoe UI", size=10, color="666666")
        r += 1
        
    # Add a thin border under reference guide
    for col in range(1, 4):
        ws.cell(r, col).border = Border(bottom=Side(style="thin", color="D3D3D3"))
    r += 2
    
    # ─── THE NEW COMPREHENSIVE GLOSSARY & METHODOLOGY SECTION ───
    ws.cell(r, 1, "📊 RESEARCH ENGINE GLOSSARY, METRICS & THESES").font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
    r += 2
    
    # Glossary Headers
    headers = ["Metric Term / Tab Name", "Practical Meaning & Underlying Investment Thesis", "Automated DB Query / Logic & Excel Formula"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(r, c_idx, h)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=Side(style="thin", color="1B365D"), bottom=Side(style="thin", color="1B365D"))
        
    ws.row_dimensions[r].height = 25
    r += 1
    
    glossary_data = [
        (
            "EMA Crossover\n(Birth of a Trend)",
            "An entry signal triggered when price crosses above its 200-day EMA.\n\n[THESIS]: Represents the birth of a new long-term uptrend. Buying here provides maximum asymmetrical risk-to-reward because the entry is close to the logical stop-loss (the EMA itself).",
            "pine_signal == 'EMA Crossover'\nSized at 7% capital allocation (~₹70,000 per position)."
        ),
        (
            "52W Breakout\n(Momentum Confirmation)",
            "A momentum signal triggered when a stock above its 200-day EMA breaks out to fresh 52-week highs.\n\n[THESIS]: Represents a confirmed breakout where institutional buying pressure clears all historical supply. High momentum outperformance.",
            "pine_signal == '52W Breakout'\nSized at 5% capital allocation (~₹50,000 per position)."
        ),
        (
            "ATH Momentum\n(Runaway Leader)",
            "An established trend signal triggered when the price is actively running near its All-Time Highs.\n\n[THESIS]: Identifies the strongest runaway leaders in the market. Ideal for high-velocity momentum trading where price experiences immediate continuation.",
            "pine_signal == 'ATH Momentum'\nSized at 4% capital allocation (~₹40,000 per position)."
        ),
        (
            "Persistence Score (30D)\n(Accumulation Footprint)",
            "The count of times a stock has appeared in the daily screener during the last 30 active trading days.\n\n[THESIS]: Represents persistent institutional buying. Random spikes appear 1-2 times, but compounders appear 10+ times as institutions accumulate.",
            "Queries the local SQLite database (`logs/backtest.db`) to sum the total number of distinct dates the ticker appeared on the screener during the last 30 active trading days.\nFormula: COUNT(DISTINCT date)"
        ),
        (
            "Consecutive Streak Days\n(Uninterrupted Trend)",
            "The number of consecutive trading days that the stock has appeared in the screener up to today.\n\n[THESIS]: High streaks indicate an exceptionally strong, uninterrupted runaway trend with heavy daily buying pressure.",
            "Checks the chronological sequence of scanner dates in SQLite starting from the most recent run date and counting backward. The streak increments for every contiguous date the ticker appears."
        ),
        (
            "Sector Velocity WoW\n(Early Sector Rotation)",
            "Compares the volume of screener signals in a sector over the last 5 trading days (W1) against the previous 5 trading days (W2).\n\n[THESIS]: Detects early sector rotation. A sudden positive surge warns you of capital flowing aggressively into a specific theme.",
            "Counts appearances of all tickers categorized by their sector. Formula:\nW1_Count = COUNT(records WHERE date IN last 5 trading days)\nW2_Count = COUNT(records WHERE date IN days -10 to -5)\nVelocity % = ((W1_Count - W2_Count) / W2_Count) * 100"
        ),
        (
            "Emerging Leaders\n(Momentum Expansion)",
            "Identifies breakout stocks experiencing a sudden momentum expansion.\n\n[THESIS]: Finds stocks with a high concentration of appearances recently but almost zero history in the preceding months. Represents new institutional entries.",
            "Finds stocks with at least 5 appearances in the last 10 trading days, but less than or equal to 3 appearances in the prior 50 trading days (i.e. trading days -60 to -10).\nFormula: COUNT(recent) >= 5 AND COUNT(historical) <= 3"
        ),
        (
            "Tab: Confluence Overlap\n(The Ultimate Overlap)",
            "This sheet displays the absolute highest conviction watchlist candidates intersecting multiple strategies.\n\n[THESIS]: Filters out market noise by only displaying stocks appearing in at least two out of three major momentum dimensions: (1) Chartink screener universe, (2) Active scanner signals, and (3) StockScans overlap list.",
            "Calculates the mathematical intersection of the three sets:\nIn_Universe + In_Scanner + In_StockScans >= 2\nSingle-dimension matches are automatically filtered out to eliminate market noise."
        ),
        (
            "Tab: Monit Non Financial\n(CHECKLIST SHEET)",
            "Your main watchlist scoring sheet for non-financial companies.\n\n[THESIS]: Keeps watchlist scoring fully formula-driven using Excel formulas. It combines automated metrics (RSI, ADX, V-stop, Persistence) with your qualitative scores (Promoter Buying, Near-Term Trigger, Headwind/Tailwind).",
            "Uses relative index-matching formulas to populate automated close prices, RSI, ADX, V-stop, and Persistence from the SQLite and Chartink universe sheets.\nFormula: INDEX-MATCH left-lookup."
        ),
        (
            "Tab: Monit Banks NBFC\n(BANKING CHECKLIST)",
            "Your main watchlist scoring sheet for banking and NBFC companies.\n\n[THESIS]: Tailored specifically for banking structures, incorporating specialized metrics like Net Interest Margins (NIM) and Credit Costs alongside technical indicators.",
            "Features banking-specific checklist criteria to ensure qualitative analyst inputs align with bank metrics."
        ),
        (
            "Tab: Scan Match tabs\n(StockScans Confluence)",
            "Displays the top 100 stocks appearing across all StockScans screeners.\n\n[THESIS]: Scores stocks qualitatively by scan volume. Stocks appearing in >20 scans receive a premium 10-point bonus, proving market-wide momentum consensus.",
            "Applies a nested Excel IF formula: `=IF(D8>20,10,IF(D8>=10,5,2))`\nScores: >20 scans = 10 pts; 10-20 scans = 5 pts; <10 scans = 2 pts."
        ),
        (
            "Leaderboard Score\n(Persistence Ranking)",
            "The primary metric used to rank stocks on the Momentum & Sector Rotation Leaderboard.\n\n[THESIS]: Institutional accumulation is a multi-week process. Combining the frequency of appearances (30D Persistence) with current trend continuation (Streak) helps identify high-probability runaway momentum leaders.",
            "Calculated dynamically in Python:\nFormula: (Persistence Score * 3) + (Consecutive Streak Days * 6)"
        ),
        (
            "Total Score\n(Confluence Ranking)",
            "The final ranking score used on the Monit scoring worksheets.\n\n[THESIS]: High-conviction watchlist candidates require a blend of technical momentum and strong fundamental catalysts. Combining Relative Strength with qualitative analyst checks ensures ranking objectivity.",
            "Calculated in Excel:\nFormula: =SUM(Relative Strength Score, [Qualitative Checklist Criterion Scores])"
        ),
        (
            "Relative Strength Rating\n(RS Rating)",
            "A percentile ranking (1 to 99) of the stock's O'Neil Weighted Score relative to the entire active universe.\n\n[THESIS]: Measures the stock's price momentum relative to all other stocks. Leaders with RS Ratings >= 90 are in the top 10% momentum class, where the strongest institutional accumulation occurs.",
            "Calculated globally across the active universe:\nFormula: Percentile Rank (1-99) of O'Neil Weighted Score"
        ),
        (
            "O'Neil Weighted Score\n(Accelerating Momentum)",
            "A momentum-based score that evaluates stock performance over the past year (53 weekly periods) divided into four quarters.\n\n[THESIS]: Based on William O'Neil's CANSLIM methodology, momentum is weighted more heavily in recent months. Double-weighting the most recent quarter (q4) highlights accelerating momentum.",
            "Calculated using weekly close prices:\nFormula: (2 * q4 + q3 + q2 + q1) / 5.0\nWhere q4 is the most recent quarter."
        ),
        (
            "Relative Strength Score\n(Sectoral Outperformance)",
            "A milestone scoring system checking relative performance against a specific industry or sector index rather than Nifty 50.\n\n[THESIS]: Stocks that cannot beat their industry peer groups are laggards. Tier-based milestone scores reward outstanding sector leaders and penalize underperformers.",
            "Calculated in Excel:\nFormula: =IF(Status=\"Underperforming\", -5, IF(Rating>=95, 15, IF(Rating>=90, 10, IF(Rating>=80, 5, 0))))"
        ),
        (
            "Calculated RSI (14)\n(Relative Strength Index)",
            "A momentum oscillator measuring the speed and change of price movements between 0 and 100.\n\n[THESIS]: Helps prevent chasing overextended rallies. Entries on low-risk pullbacks (RSI < 40) are preferred over buying vertical extensions (RSI > 70).",
            "Calculated using the standard Wilders smoothing method over a 14-period daily close price window."
        ),
        (
            "Calculated ADX (14)\n(Average Directional Index)",
            "A technical indicator measuring the overall strength of a trend on a scale from 0 to 100, independent of trend direction.\n\n[THESIS]: Distinguishes between choppy ranges and active breakout trends. An ADX value > 25 confirms a strong, sustainable trending environment.",
            "Calculated over a 14-period window using Daily High, Low, and Close prices."
        ),
        (
            "Calculated V-stop Line\n(Volatility Stop)",
            "A volatility-based trailing stop-loss boundary derived from the Average True Range (ATR).\n\n[THESIS]: Serves as a dynamic trailing stop-loss. A price drop below this line indicates a breakdown of the trend's volatility boundary, signaling a potential exit.",
            "Calculated using the stock's true range adjusted by a volatility multiplier. Plotting provides a trailing stop reference."
        ),
        (
            "Calculated Near-term Trigger\n(Breakout Confirmation)",
            "Identifies specific consolidation contractions or sudden volume breakout triggers based on recent prices.\n\n[THESIS]: Pinpoints the exact day of momentum expansion to trigger active buy/sell execution.",
            "Checks recent volume spikes and tight price ranges to flag patterns like 'Volume Surge' or 'Range Breakout'."
        ),
        (
            "Weekly Tracking Highlighting\n(Additions / Removals)",
            "Visual highlights indicating weekly changes in the StockScans 100 (Scan Match) list.\n\n[THESIS]: Helps track the weekly turnover. Newly added candidates indicate fresh institutional attention, while removed candidates represent names losing relative strength.",
            "Compares daily lists against Monday's baseline database:\n- Added stocks: Highlighted in Soft Blue.\n- Removed stocks: Listed in the bottom section of Scan Match sheets."
        )
    ]
    
    # Border styles
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0")
    )
    
    for term, meaning, formula in glossary_data:
        # Term (A)
        c_term = ws.cell(r, 1, term)
        c_term.font = Font(name="Segoe UI", size=10, bold=True, color="1B365D")
        c_term.alignment = Alignment(vertical="top", wrap_text=True)
        c_term.border = thin_border
        c_term.fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
        
        # Meaning (B)
        c_meaning = ws.cell(r, 2, meaning)
        c_meaning.font = Font(name="Segoe UI", size=10, color="333333")
        c_meaning.alignment = Alignment(vertical="top", wrap_text=True)
        c_meaning.border = thin_border
        
        # Formula/Logic (C)
        c_formula = ws.cell(r, 3, formula)
        c_formula.font = Font(name="Segoe UI", size=9.5, color="444444", italic=True)
        c_formula.alignment = Alignment(vertical="top", wrap_text=True)
        c_formula.border = thin_border
        
        # Set line height adaptively
        max_lines = max(len(term.split('\n')), len(meaning.split('\n')), len(formula.split('\n')))
        ws.row_dimensions[r].height = max(55, max_lines * 16 + 12)
        
        r += 1
        
    # Auto-adjust column sizes for the sheet
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 65
    
    # Freeze pane is not needed for a static guide
    ws.freeze_panes = None


def write_universe_sheet(ws, universe: pd.DataFrame) -> None:
    cols = [
        "symbol",
        "company",
        "sector",
        "industry",
        "marketcap_bucket",
        "marketcap_cr",
        "close",
        "today_return_pct",
        "volume",
        "bse_code",
    ]
    write_table(ws, "A1", cols, universe[cols].itertuples(index=False, name=None))
    style_basic_table(ws, 1, len(universe) + 1, len(cols))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def load_scanner_signals_dict() -> dict[str, dict]:
    """Read full scanner signals as a lookup dictionary symbol -> signal metrics."""
    from datetime import date
    csv_path = Path("logs") / f"signals_{date.today()}.csv"
    if not csv_path.exists():
        csv_files = sorted(Path("logs").glob("signals_*.csv"))
        if csv_files:
            csv_path = csv_files[-1]
        else:
            return {}
    try:
        df = pd.read_csv(csv_path)
        if "ticker" in df.columns:
            signals_dict = {}
            for _, row in df.iterrows():
                ticker = str(row["ticker"]).strip()
                signals_dict[ticker] = {
                    "entry": str(row.get("entry", "No Signal")),
                    "score": float(row.get("score", 0.0)),
                    "alloc_pct": str(row.get("alloc_pct", "0%")),
                    "alloc_inr": float(row.get("alloc_inr", 0.0)),
                    "qty": int(row.get("qty", 0)),
                }
            return signals_dict
    except Exception:
        pass
    return {}


def load_backtest_recurrence_dict() -> dict[str, int]:
    """Read local backtest DB and count occurrences of each symbol in the last 30 trading days."""
    try:
        df = load_backtest_df()
        if df.empty:
            return {}
        unique_dates = sorted(df["parsed_date"].unique())
        last_30_dates = unique_dates[-30:] if len(unique_dates) >= 30 else unique_dates
        df_30 = df[df["parsed_date"].isin(last_30_dates)]
        counts = df_30.groupby("symbol")["date"].nunique().to_dict()
        return counts
    except Exception as e:
        print(f"⚠️ Error loading backtest recurrence: {e}")
        return {}


def write_scoring_sheet(
    ws, 
    data: pd.DataFrame, 
    criteria: list[Criterion], 
    source_sheet_name: str,
    existing_inputs: dict[str, dict[str, str]],
    yfinance_data: dict[str, dict],
    global_rs: dict[str, dict],
    include_portfolio: bool = False,
    added_symbols: set[str] = None,
    removed_symbols: dict[str, str] = None
) -> dict[str, float]:
    signals_dict = load_scanner_signals_dict() if include_portfolio else {}
    if include_portfolio:
        meta_headers = [
            "Rank",
            "Total Score",
            "Entry",
            "Entry Score",
            "Allocation %",
            "Allocation (₹)",
            "Quantity",
            "Symbol",
            "Company",
            "Sector",
            "Industry",
            "Market Cap Bucket",
            "Market Cap Cr",
            "Close",
            "Today Return %",
            "Volume",
            "Relative Strength Rating",
            "Relative Strength Status",
            "Relative Strength Score",
            "Calculated RSI (14)",
            "Calculated ADX (14)",
            "Calculated V-stop Line",
            "Calculated Near-term Trigger",
            "Sector Benchmark",
            "Benchmark 1Y Return %",
            "Stock 1Y Return %",
            "Relative Strength 1Y (%)",
            "O'Neil Weighted Score"
        ]
    else:
        meta_headers = [
            "Rank",
            "Total Score",
            "Symbol",
            "Company",
            "Sector",
            "Industry",
            "Market Cap Bucket",
            "Market Cap Cr",
            "Close",
            "Today Return %",
            "Volume",
            "Relative Strength Rating",
            "Relative Strength Status",
            "Relative Strength Score",
            "Calculated RSI (14)",
            "Calculated ADX (14)",
            "Calculated V-stop Line",
            "Calculated Near-term Trigger",
            "Sector Benchmark",
            "Benchmark 1Y Return %",
            "Stock 1Y Return %",
            "Relative Strength 1Y (%)",
            "O'Neil Weighted Score"
        ]
    headers = meta_headers[:]
    for criterion in criteria:
        headers.append(criterion.label)
        headers.append(f"{criterion.label} Score")
    ws.append(headers)

    max_row = len(data) + 1
    score_cols: list[int] = []
    
    # 1. Pre-calculate the total score for each stock in Python so we can sort them!
    enriched_rows = []
    for _, row in data.iterrows():
        company_name = row["company"]
        symbol = row["symbol"]
        matched_company = find_matching_company(company_name, existing_inputs.keys())
        
        details = yfinance_data.get(symbol, {})
        is_bank = source_sheet_name == "Banks & NBFC"
        
        # Retrieve pre-calculated RS metrics from global dict
        rs_metrics = global_rs.get(symbol, {
            "bench_name": "",
            "stock_1y_ret": None,
            "bench_1y_ret": None,
            "rs_spread": None,
            "weighted_score": None,
            "rs_rating": "",
            "rs_status": ""
        })
        
        row_inputs = {}
        total_score = 0.0
        for criterion in criteria:
            val = None
            # Copy existing inputs from template
            if matched_company and criterion.label in existing_inputs[matched_company]:
                val = existing_inputs[matched_company][criterion.label]
            # Otherwise use automated internet calculation
            if val is None and details:
                val = get_automated_input_value(criterion.label, row, details, is_bank)
            
            row_inputs[criterion.label] = val
            if val is not None:
                total_score += evaluate_excel_if_formula(criterion.score_formula, val)
                
        # Calculate Relative Strength Score based on Option 1 (Milestone Tiers)
        rs_score_val = 0.0
        rs_rating = rs_metrics.get("rs_rating")
        rs_status = rs_metrics.get("rs_status", "")
        if rs_status == "Underperforming":
            rs_score_val = -5.0
        elif rs_rating != "" and rs_rating is not None:
            try:
                rating_val = int(rs_rating)
                if rating_val >= 95:
                    rs_score_val = 15.0
                elif rating_val >= 90:
                    rs_score_val = 10.0
                elif rating_val >= 80:
                    rs_score_val = 5.0
            except ValueError:
                pass
                
        total_score += rs_score_val
                
        enriched_rows.append({
            "row": row,
            "inputs": row_inputs,
            "total_score": total_score,
            "rs_metrics": rs_metrics
        })
        
    # 2. Sort the rows by total_score in descending order!
    enriched_rows = sorted(enriched_rows, key=lambda x: x["total_score"], reverse=True)

    # Build dynamic column map dictionary to bypass hardcoded column index shifts
    col_map = {name: col_idx for col_idx, name in enumerate(meta_headers, start=1)}

    # 3. Write metadata and automated/prepopulated inputs in sorted order
    for idx, item in enumerate(enriched_rows, start=2):
        row = item["row"]
        row_inputs = item["inputs"]
        symbol = row["symbol"]
        details = yfinance_data.get(symbol, {})
        rs_metrics = item["rs_metrics"]
        rs_status = rs_metrics.get("rs_status", "")
        
        # Calculate raw technical values
        rsi_val = ""
        adx_val = ""
        vstop_val = ""
        trigger_val = ""
        if details:
            close_prices = details.get("close", None)
            high_prices = details.get("high", None)
            low_prices = details.get("low", None)
            volume_prices = details.get("volume", None)
            
            rsi_raw = calculate_rsi(close_prices)
            adx_raw = calculate_adx(high_prices, low_prices, close_prices)
            vstop_raw = calculate_vstop_price(high_prices, low_prices, close_prices)
            trigger_raw = calculate_near_term_trigger_val(close_prices, volume_prices)
            
            rsi_val = round(rsi_raw, 2) if rsi_raw else ""
            adx_val = round(adx_raw, 2) if adx_raw else ""
            vstop_val = round(vstop_raw, 2) if vstop_raw else ""
            trigger_val = trigger_raw

        # Rank and Total Score columns are formula-driven, set placeholders
        ws.cell(idx, 1, "")
        ws.cell(idx, 2, "")

        if include_portfolio:
            sig = signals_dict.get(symbol, {
                "entry": "No Signal",
                "score": 0.0,
                "alloc_pct": "0%",
                "alloc_inr": 0.0,
                "qty": 0
            })
            ws.cell(idx, col_map["Entry"], sig["entry"])
            ws.cell(idx, col_map["Entry Score"], sig["score"])
            ws.cell(idx, col_map["Allocation %"], sig["alloc_pct"])
            ws.cell(idx, col_map["Allocation (₹)"], sig["alloc_inr"])
            ws.cell(idx, col_map["Quantity"], sig["qty"])

        c_sym = ws.cell(idx, col_map["Symbol"], row["symbol"])
        c_comp = ws.cell(idx, col_map["Company"], row["company"])
        if added_symbols and row["symbol"] in added_symbols:
            added_fill = PatternFill("solid", fgColor="D9E1F2") # Soft Blue
            c_sym.fill = added_fill
            c_comp.fill = added_fill
        elif rs_status == "Underperforming":
            under_fill = PatternFill("solid", fgColor="F2DCDB") # Soft Red
            c_sym.fill = under_fill
            c_comp.fill = under_fill
        ws.cell(idx, col_map["Sector"], row.get("sector") or "")
        ws.cell(idx, col_map["Industry"], row.get("industry") or "")
        ws.cell(idx, col_map["Market Cap Bucket"], row.get("marketcap_bucket") or "")
        ws.cell(idx, col_map["Market Cap Cr"], row.get("marketcap_cr") or 0.0)
        ws.cell(idx, col_map["Close"], row.get("close") or 0.0)
        ws.cell(idx, col_map["Today Return %"], row.get("today_return_pct") or 0.0)
        ws.cell(idx, col_map["Volume"], row.get("volume") or 0.0)
        
        # Write technical indicator columns
        ws.cell(idx, col_map["Calculated RSI (14)"], rsi_val)
        ws.cell(idx, col_map["Calculated ADX (14)"], adx_val)
        ws.cell(idx, col_map["Calculated V-stop Line"], vstop_val)
        ws.cell(idx, col_map["Calculated Near-term Trigger"], trigger_val)
        
        # Write sectoral Nifty benchmark relative strength columns
        
        ws.cell(idx, col_map["Sector Benchmark"], rs_metrics["bench_name"] or "")
        ws.cell(idx, col_map["Benchmark 1Y Return %"], rs_metrics["bench_1y_ret"] if rs_metrics["bench_1y_ret"] is not None else "")
        ws.cell(idx, col_map["Stock 1Y Return %"], rs_metrics["stock_1y_ret"] if rs_metrics["stock_1y_ret"] is not None else "")
        ws.cell(idx, col_map["Relative Strength 1Y (%)"], rs_metrics["rs_spread"] if rs_metrics["rs_spread"] is not None else "")
        ws.cell(idx, col_map["O'Neil Weighted Score"], rs_metrics["weighted_score"] if rs_metrics["weighted_score"] is not None else "")
        ws.cell(idx, col_map["Relative Strength Rating"], rs_metrics["rs_rating"])
        ws.cell(idx, col_map["Relative Strength Status"], rs_status)
        
        # Populate dynamic Excel formula for Option 1 Relative Strength Score
        rating_letter = get_column_letter(col_map["Relative Strength Rating"])
        status_letter = get_column_letter(col_map["Relative Strength Status"])
        rs_score_formula = f'=IF({status_letter}{idx}="Underperforming",-5,IF({rating_letter}{idx}>=95,15,IF({rating_letter}{idx}>=90,10,IF({rating_letter}{idx}>=80,5,0))))'
        ws.cell(idx, col_map["Relative Strength Score"], rs_score_formula)

        # Write qualitative input columns
        input_start_col = len(meta_headers) + 1
        for i, criterion in enumerate(criteria):
            input_col = input_start_col + i * 2
            val = row_inputs.get(criterion.label)
            if val is not None:
                ws.cell(idx, input_col, val)

    input_start_col = len(meta_headers) + 1
    for i, criterion in enumerate(criteria):
        input_col = input_start_col + i * 2
        score_col = input_col + 1
        score_cols.append(score_col)
        input_letter = get_column_letter(input_col)
        score_letter = get_column_letter(score_col)
        ws.cell(1, input_col).comment = None

    if max_row >= 2:
        for i, criterion in enumerate(criteria):
            input_col = input_start_col + i * 2
            score_col = input_col + 1
            input_letter = get_column_letter(input_col)
            if criterion.validation_formula:
                dv = DataValidation(type="list", formula1=criterion.validation_formula, allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(f"{input_letter}2:{input_letter}{max_row}")
            for row_idx in range(2, max_row + 1):
                translated = translate_score_formula(
                    criterion.score_formula,
                    criterion.source_row,
                    f"{input_letter}{row_idx}",
                )
                ws.cell(row_idx, score_col, clean_excel_formula(translated))

        rs_score_col_letter = get_column_letter(col_map["Relative Strength Score"])
        for row_idx in range(2, max_row + 1):
            refs = [f"{rs_score_col_letter}{row_idx}"] + [
                f"{get_column_letter(col)}{row_idx}"
                for col in score_cols
            ]
            ws.cell(row_idx, 2, f"=SUM({','.join(refs)})")
            ws.cell(row_idx, 1, f'=IF(B{row_idx}=0,"",RANK(B{row_idx},$B$2:$B${max_row},0))')

        style_basic_table(ws, 1, max_row, len(headers), include_portfolio=include_portfolio)
        volume_col_idx = col_map["Volume"]
        ws.freeze_panes = get_column_letter(volume_col_idx + 1) + "2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max_row}"
        # Highlight Total Score > 250 with a premium highlight (Light Green fill, Dark Green font)
        from openpyxl.formatting.rule import CellIsRule
        high_score_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        high_score_font = Font(color="006100", bold=True)
        ws.conditional_formatting.add(
            f"B2:B{max_row}",
            CellIsRule(operator='greaterThan', formula=['250'], stopIfTrue=True, fill=high_score_fill, font=high_score_font)
        )

        add_score_conditional_format(ws, 2, 2, max_row)
        for col in score_cols:
            add_score_conditional_format(ws, col, 2, max_row)
        
        # Apply custom color coding to Relative Strength Rating
        rs_rating_col = col_map["Relative Strength Rating"]
        add_rs_rating_conditional_formatting(ws, rs_rating_col, 2, max_row)
        
        # Apply red color coding for Underperforming Status and Score
        from openpyxl.formatting.rule import CellIsRule
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006", bold=True)
        
        status_col_letter = get_column_letter(col_map["Relative Strength Status"])
        ws.conditional_formatting.add(
            f"{status_col_letter}2:{status_col_letter}{max_row}",
            CellIsRule(operator='equal', formula=['"Underperforming"'], stopIfTrue=True, fill=red_fill, font=red_font)
        )
        
        score_col_letter = get_column_letter(col_map["Relative Strength Score"])
        ws.conditional_formatting.add(
            f"{score_col_letter}2:{score_col_letter}{max_row}",
            CellIsRule(operator='equal', formula=['-5'], stopIfTrue=True, fill=red_fill, font=red_font)
        )

        # Apply conditional formatting for Calculated V-stop Line (Close >= V-stop is green, Close < V-stop is red)
        from openpyxl.formatting.rule import CellIsRule
        close_letter = get_column_letter(col_map["Close"])
        vstop_letter = get_column_letter(col_map["Calculated V-stop Line"])
        
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100", bold=True)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006", bold=True)
        
        # Rule 1: Blank check to stop evaluation for empty V-stops (so they don't turn green/red)
        ws.conditional_formatting.add(
            f"{vstop_letter}2:{vstop_letter}{max_row}",
            CellIsRule(operator='equal', formula=['""'], stopIfTrue=True)
        )
        # Rule 2: Green if Vstop <= Close
        ws.conditional_formatting.add(
            f"{vstop_letter}2:{vstop_letter}{max_row}",
            CellIsRule(operator='lessThanOrEqual', formula=[f"={close_letter}2"], stopIfTrue=True, fill=green_fill, font=green_font)
        )
        # Rule 3: Red if Vstop > Close
        ws.conditional_formatting.add(
            f"{vstop_letter}2:{vstop_letter}{max_row}",
            CellIsRule(operator='greaterThan', formula=[f"={close_letter}2"], stopIfTrue=True, fill=red_fill, font=red_font)
        )

        # Write weekly tracking removals for Scan Match sheets
        if ws.title.startswith("Scan Match") and removed_symbols:
            start_removed_row = max_row + 4
            
            title_cell = ws.cell(start_removed_row, 1, "❌ REMOVED SCAN MATCH STOCKS (THIS WEEK)")
            title_cell.font = Font(name="Calibri", size=12, bold=True, color="C00000")
            
            rem_headers = ["Symbol", "Company Name", "Sector", "Date of Removal"]
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="595959")
            thin_border = Border(
                left=Side(style="thin", color="D9E2F3"),
                right=Side(style="thin", color="D9E2F3"),
                top=Side(style="thin", color="D9E2F3"),
                bottom=Side(style="thin", color="D9E2F3")
            )
            
            for col_idx, h in enumerate(rem_headers, start=1):
                c = ws.cell(start_removed_row + 1, col_idx, h)
                c.font = header_font
                c.fill = header_fill
                c.border = thin_border
                c.alignment = Alignment(horizontal="center", vertical="center")
                
            r_row = start_removed_row + 2
            is_banking_sheet = source_sheet_name == "Banks & NBFC"
            
            for sym, rem_date in sorted(removed_symbols.items()):
                # Classify symbol to verify if it belongs to this sheet
                is_fin = False
                details = yfinance_data.get(sym, {})
                info = details.get("info", {})
                
                dummy_row = pd.Series({
                    "symbol": sym,
                    "company": info.get("longName", sym),
                    "sector": info.get("sector", "unknown"),
                    "industry": info.get("industry", "unknown")
                })
                is_fin = is_financial_company(dummy_row)
                
                if (is_banking_sheet and not is_fin) or (not is_banking_sheet and is_fin):
                    continue
                    
                comp_name = info.get("longName", sym)
                sector = info.get("sector", "unknown")
                
                row_fill = PatternFill("solid", fgColor="F9EBEA") # Soft light red/pink fill for removed status
                
                c_sym = ws.cell(r_row, 1, sym)
                c_sym.font = Font(bold=True)
                c_sym.border = thin_border
                c_sym.fill = row_fill
                c_sym.alignment = Alignment(horizontal="center")
                
                c_name = ws.cell(r_row, 2, comp_name)
                c_name.border = thin_border
                c_name.fill = row_fill
                
                c_sect = ws.cell(r_row, 3, sector)
                c_sect.border = thin_border
                c_sect.fill = row_fill
                
                c_date = ws.cell(r_row, 4, rem_date)
                c_date.border = thin_border
                c_date.fill = row_fill
                c_date.alignment = Alignment(horizontal="center")
                
                r_row += 1
    else:
        style_basic_table(ws, 1, 1, len(headers), include_portfolio=include_portfolio)

    ws["A1"].comment = None
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = f"Rank ({source_sheet_name})"
    
    scores_dict = {}
    for item in enriched_rows:
        scores_dict[item["row"]["symbol"]] = item["total_score"]
    return scores_dict


def write_formula_map(
    ws, non_financial: list[Criterion], banks: list[Criterion]
) -> None:
    ws.append(["Model", "Section", "Criterion", "Question", "Source Row", "Score Formula", "Dropdown Options"])
    for model, criteria in [("Non Financial", non_financial), ("Banks & NBFC", banks)]:
        for c in criteria:
            ws.append([
                model,
                c.section,
                c.label,
                c.question,
                c.source_row,
                "'" + c.score_formula,
                c.validation_formula,
            ])
    style_basic_table(ws, 1, ws.max_row, 7)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_confluence_sheet(
    ws,
    union_symbols: set[str],
    universe: pd.DataFrame,
    scanner_symbols: list[str],
    scan_matched_symbols: list[str],
    yfinance_data: dict[str, dict],
    all_scores: dict[str, float],
    scans_dict: dict[str, int],
    signals_dict: dict[str, dict],
    industry_dict: dict[str, str],
    backtest_recurrence_dict: dict[str, int]
) -> None:
    """Write the Confluence Overlap sheet with dynamic left-lookups for scoring and enriched industries."""
    headers = [
        "Confluence Rank",
        "Total Score",
        "Relative Strength Rating",
        "Symbol",
        "Company Name",
        "Industry",
        "Common Count",
        "Scanner Signal",
        "Chartink Universe",
        "Scan Matches Count",
        "Backtest Recurrence (30D)"
    ]
    ws.append(headers)
    
    set_universe = set(universe["symbol"].dropna().unique())
    set_scanner = set(scanner_symbols)
    set_scan_match = set(scan_matched_symbols)
    
    rows = []
    for sym in union_symbols:
        in_univ = sym in set_universe
        in_scan = sym in set_scanner
        in_match = sym in set_scan_match
        
        overlap_count = int(in_univ) + int(in_scan) + int(in_match)
        
        # Skip single match stocks (Common Count < 2)
        if overlap_count < 2:
            continue
            
        company_name = sym
        if in_univ:
            company_name = universe[universe["symbol"] == sym].iloc[0]["company"]
        elif in_match and sym in yfinance_data:
            company_name = yfinance_data[sym].get("info", {}).get("longName", sym)
            
        py_score = all_scores.get(sym, 0.0)
        
        rows.append({
            "symbol": sym,
            "company": company_name,
            "industry": industry_dict.get(sym, "unknown"),
            "overlap_count": overlap_count,
            "in_univ": "Yes" if in_univ else "No",
            "in_scan": signals_dict.get(sym, {}).get("entry", "Yes") if in_scan else "No",
            "scans_count": scans_dict.get(sym, 0),
            "py_score": py_score,
            "recurrence": backtest_recurrence_dict.get(sym, 0)
        })
        
    # Sort: 1st by overlap_count (desc), 2nd by py_score (desc)
    sorted_rows = sorted(rows, key=lambda x: (x["overlap_count"], x["py_score"]), reverse=True)
    
    max_row = len(sorted_rows) + 1
    for idx, item in enumerate(sorted_rows, start=2):
        sym = item["symbol"]
        ws.cell(idx, 4, sym)
        ws.cell(idx, 5, item["company"])
        ws.cell(idx, 6, item["industry"])
        ws.cell(idx, 7, item["overlap_count"])
        ws.cell(idx, 8, item["in_scan"])
        ws.cell(idx, 9, item["in_univ"])
        ws.cell(idx, 10, item["scans_count"])
        ws.cell(idx, 11, item["recurrence"])
        
        # INDEX-MATCH lookup for scores
        ws.cell(idx, 2, f"=IFERROR(INDEX('Monit Non Financial'!$B$2:$B$400, MATCH(D{idx}, 'Monit Non Financial'!$C$2:$C$400, 0)), IFERROR(INDEX('Monit Banks NBFC'!$B$2:$B$100, MATCH(D{idx}, 'Monit Banks NBFC'!$C$2:$C$100, 0)), 0))")
        
        # INDEX-MATCH lookup for Relative Strength Rating (col L in Monit sheets)
        ws.cell(idx, 3, f"=IFERROR(INDEX('Monit Non Financial'!$L$2:$L$400, MATCH(D{idx}, 'Monit Non Financial'!$C$2:$C$400, 0)), IFERROR(INDEX('Monit Banks NBFC'!$L$2:$L$100, MATCH(D{idx}, 'Monit Banks NBFC'!$C$2:$C$100, 0)), \"\"))")
        
        # Confluence Rank
        ws.cell(idx, 1, f'=IF(B{idx}=0,"",ROW()-1)')
        
    style_basic_table(ws, 1, max_row, len(headers))
    ws.freeze_panes = "F2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    add_score_conditional_format(ws, 7, 2, max_row)
    add_rs_rating_conditional_formatting(ws, 3, 2, max_row)


def translate_score_formula(formula: str, source_row: int, target_input_cell: str) -> str:
    """Move one SOIC score formula from the template input cell to a row input cell."""
    pattern = re.compile(rf"(?<![A-Z0-9_])\$?D\$?{source_row}(?![0-9])")
    return pattern.sub(target_input_cell, formula)


def write_table(ws, start_cell: str, headers: list[str], rows: Iterable[tuple]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(list(row))


def style_basic_table(ws, header_row: int, max_row: int, max_col: int, include_portfolio: bool = False) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    for col in range(1, max_col + 1):
        cell = ws.cell(header_row, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        width = min(max(len(str(cell.value or "")) + 4, 12), 34)
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(header_row + 1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
    # Dynamically format float/numeric columns
    fmt_cols = []
    for col in range(1, max_col + 1):
        header_val = ws.cell(header_row, col).value
        if header_val in (
            "Total Score", "Entry Score", "Allocation %", "Allocation (₹)", "Quantity",
            "Market Cap Cr", "Close", "Today Return %", "Volume",
            "Calculated RSI (14)", "Calculated ADX (14)", "Calculated V-stop Line",
            "Benchmark 1Y Return %", "Stock 1Y Return %", "Relative Strength 1Y (%)", "O'Neil Weighted Score", "Relative Strength Score"
        ):
            fmt_cols.append(col)
            
    for col in fmt_cols:
        if col <= max_col:
            for row in range(header_row + 1, max_row + 1):
                ws.cell(row, col).number_format = "#,##0.00"


def add_score_conditional_format(ws, col: int, start_row: int, end_row: int) -> None:
    letter = get_column_letter(col)
    ws.conditional_formatting.add(
        f"{letter}{start_row}:{letter}{end_row}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )


def send_cloud_alerts(excel_path: Path, universe_len: int) -> None:
    """Send daily Excel workbook and summary report via Email and Telegram."""
    global STOCKSCANS_STATUS, CONFLUENCE_EMERGING_REPORT, CONFLUENCE_EMERGING_HTML
    import os
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication
    from datetime import date
    
    # 1. Fetch credentials from environment
    gmail_user = os.environ.get("GMAIL_USER", "")
    gmail_app_pass = os.environ.get("GMAIL_APP_PASS", "")
    alert_email = os.environ.get("ALERT_EMAIL", gmail_user)
    
    tg_token = os.environ.get("TELEGRAM_TOKEN", "")
    tg_chat = os.environ.get("TELEGRAM_CHAT_ID", "")
    
    today_str = date.today().strftime("%d %b %Y")
    filename = excel_path.name

    # Check StockScans API Connection Status for warnings
    warning_text = ""
    warning_html = ""
    warning_tg = ""
    warning_wa = ""
    
    if not STOCKSCANS_STATUS.get("fetched_live", False):
        reason = STOCKSCANS_STATUS.get("message", "Cookie failed or expired.")
        warning_text = f"⚠️ [STOCKSCANS ALERT]: The live StockScans.in connection failed or cookie has EXPIRED.\n" \
                       f"Reason: {reason}\n" \
                       f"The system fell back to the committed cache. Please update the 'STOCKSCANS_COOKIE' repository secret on GitHub settings to restore live confluences!\n" \
                       f"────────────────────────────────────────────────────────────\n\n"
                       
        warning_html = f"""
        <div style="background-color:#fff2f2;border-left:4px solid #ff4d4d;padding:12px;margin-bottom:15px;border-radius:4px;font-family:sans-serif">
          <b style="color:#cc0000">⚠️ Live StockScans Connection Failed / Cookie Expired</b><br>
          The automated cloud runner failed to establish an authenticated connection to StockScans.in (Reason: {reason}).<br>
          The system successfully fell back to your committed offline cache file so sheets compiled successfully.<br>
          <span style="font-size:12px;color:#555">👉 Please inspect headers on StockScans, copy your new cookie, and update the <b>STOCKSCANS_COOKIE</b> secret in your GitHub repository settings to restore live confluences.</span>
        </div>
        """
        
        warning_tg = "⚠️ *StockScans Cookie Expired/Invalid* — fell back to cache. Please update the `STOCKSCANS_COOKIE` secret on GitHub settings!\n\n"
        
        warning_wa = "⚠️ *StockScans Cookie Expired/Invalid* — fell back to offline cache. Please update the `STOCKSCANS_COOKIE` secret on GitHub settings!\n\n"
    
    # 2. Attempt to load today's screener signals from the CSV logged by scanner.py
    signals_file = Path(f"logs/signals_{date.today()}.csv")
    signals_df = pd.DataFrame()
    if signals_file.exists():
        try:
            signals_df = pd.read_csv(signals_file)
        except Exception:
            pass

    has_signals = not signals_df.empty and len(signals_df) > 0
    ENTRY_EMOJI = {"EMA Crossover": "🟢", "52W Breakout": "🟡", "ATH Momentum": "🔵"}
    TOTAL_CAPITAL = 1_000_000

    # 3. Format unified Plain Text and HTML signal summaries
    if has_signals:
        n_cross = len(signals_df[signals_df["entry"] == "EMA Crossover"])
        n_new   = len(signals_df[signals_df["entry"] == "52W Breakout"])
        n_run   = len(signals_df[signals_df["entry"] == "ATH Momentum"])
        total_deployed = signals_df["alloc_inr"].sum()
        
        # Build plain text watchlist lines
        sig_lines = []
        for _, row in signals_df.iterrows():
            e = ENTRY_EMOJI.get(row["entry"], "⚪")
            sig_lines.append(
                f"{e} {row['ticker']:12s} | ₹{row['close']:>8.2f} | "
                f"RSI {row['rsi']:4.1f} | {row['pct_ath']:4.1f}% from ATH | "
                f"{row['entry']:13s} | Alloc ₹{row['alloc_inr']:>7,} ({row['qty']} qty)"
            )
        signals_plain_list = "\n".join(sig_lines)
        
        plain_text = warning_text + \
                     f"🏆 Daily Monit Multibagger Watchlist — {today_str}\n\n" \
                     f"The automated ranking pipeline has completed successfully.\n" \
                     f"• Scored & Ranked: {universe_len} companies\n" \
                     f"• Local SQLite database updated and synced.\n\n" \
                     f"📊 Today's Screener Signals:\n" \
                     f"🟢 EMA Crossover: {n_cross}  🟡 52W Breakout: {n_new}  🔵 ATH Momentum: {n_run}\n" \
                     f"Capital deployed: ₹{total_deployed:,.0f} / ₹{TOTAL_CAPITAL:,.0f}\n" \
                     f"────────────────────────────────────────────────────────────\n" \
                     f"{signals_plain_list}\n" \
                     f"────────────────────────────────────────────────────────────\n\n"
                     
        if CONFLUENCE_EMERGING_REPORT:
            plain_text += f"{CONFLUENCE_EMERGING_REPORT}\n" \
                          f"────────────────────────────────────────────────────────────\n\n"
                          
        plain_text += f"Your premium daily Excel workbook is attached to this email.\n\n" \
                      f"Best regards,\nMonit Research Desk"
        
        # Build HTML table for Gmail
        rows_html = ""
        for _, row in signals_df.iterrows():
            e = ENTRY_EMOJI.get(row["entry"], "⚪")
            bg = {"EMA Crossover": "#e6ffe6", "52W Breakout": "#fffbe6", "ATH Momentum": "#e6f0ff"}.get(row["entry"], "#fff")
            rows_html += f"""
            <tr style="background:{bg}">
              <td style="padding:8px;font-weight:bold">{e} {row['ticker']}</td>
              <td style="padding:8px;text-align:right">₹{row['close']:,.2f}</td>
              <td style="padding:8px;text-align:right">{row['rsi']}</td>
              <td style="padding:8px;text-align:right">{row['pct_ath']}%</td>
              <td style="padding:8px;font-weight:bold;color:{'#1a7a1a' if row['entry']=='EMA Crossover' else '#7a6a00' if row['entry']=='52W Breakout' else '#003e7a'}">{row['entry']}</td>
              <td style="padding:8px;text-align:right">₹{row['alloc_inr']:,}</td>
              <td style="padding:8px;text-align:right">{row['qty']}</td>
            </tr>"""

        html_text = f"""
        <html><body style="font-family:sans-serif;max-width:850px;margin:auto">
        <h2 style="color:#1B365D">🏆 Daily Monit Multibagger Watchlist — {today_str}</h2>
        {warning_html}
        <p>The automated ranking pipeline has completed successfully today.</p>
        <ul>
          <li><b>Scored & Ranked:</b> {universe_len} companies</li>
          <li><b>Local Database:</b> Updated and synced (screener history ledger + rollup analytics)</li>
        </ul>
        <br>
        <h3 style="color:#1a1a2e">📊 Screener Signals Generated Today:</h3>
        <p>
          <b style="color:#1a7a1a">🟢 EMA Crossover: {n_cross}</b> &nbsp;
          <b style="color:#7a6a00">🟡 52W Breakout: {n_new}</b> &nbsp;
          <b style="color:#003e7a">🔵 ATH Momentum: {n_run}</b> &nbsp;&nbsp;|&nbsp;&nbsp;
          Capital deployed: <b>₹{total_deployed:,.0f}</b> of ₹{TOTAL_CAPITAL:,.0f}
        </p>
        <table border="1" cellspacing="0" style="border-collapse:collapse;width:100%">
          <thead style="background:#1a1a2e;color:white">
            <tr>
              <th style="padding:8px">Stock</th>
              <th style="padding:8px">Close</th>
              <th style="padding:8px">RSI</th>
              <th style="padding:8px">% from ATH</th>
              <th style="padding:8px">Entry</th>
              <th style="padding:8px">Alloc (₹)</th>
              <th style="padding:8px">Qty</th>
            </tr>
          </thead>
          <tbody>{rows_html}</tbody>
        </table>
        
        {CONFLUENCE_EMERGING_HTML}
        
        <br>
        <p>Your premium daily watchlist Excel workbook <b>{filename}</b> has been compiled and is attached below.</p>
        <br>
        <p style="color:#555;font-size:12px">⚠️ This is an automated scan alert. Always verify charts on TradingView before investing.</p>
        </body></html>
        """
    else:
        plain_text = warning_text + \
                     f"🏆 Daily Monit Multibagger Watchlist — {today_str}\n\n" \
                     f"The automated ranking pipeline has completed successfully.\n" \
                     f"• Scored & Ranked: {universe_len} companies\n" \
                     f"• Local SQLite database updated and synced.\n\n" \
                     f"ℹ️ No active screener signals were generated by the Pine Script engine today.\n\n"
                     
        if CONFLUENCE_EMERGING_REPORT:
            plain_text += f"{CONFLUENCE_EMERGING_REPORT}\n" \
                          f"────────────────────────────────────────────────────────────\n\n"
                          
        plain_text += f"Your daily premium Excel workbook is attached to this email.\n\n" \
                      f"Best regards,\nMonit Research Desk"
                     
        html_text = f"""
        <html><body style="font-family:sans-serif;max-width:850px;margin:auto">
        <h2 style="color:#1B365D">🏆 Daily Monit Multibagger Watchlist — {today_str}</h2>
        {warning_html}
        <p>The automated ranking pipeline has completed successfully today.</p>
        <ul>
          <li><b>Scored & Ranked:</b> {universe_len} companies</li>
          <li><b>Local Database:</b> Updated and synced (screener history ledger + rollup analytics)</li>
        </ul>
        <br>
        <p><b>ℹ️ No active screener signals were generated by the Pine Script engine today.</b></p>
        
        {CONFLUENCE_EMERGING_HTML}
        
        <br>
        <p>Your premium watchlist Excel workbook <b>{filename}</b> has been compiled and is attached below.</p>
        <br>
        <p style="color:#555;font-size:12px">⚠️ This is an automated scan alert. Always verify charts on TradingView before investing.</p>
        </body></html>
        """

    # 4. Send the Unified Email with Workbook Attached
    if gmail_user and gmail_app_pass:
        try:
            recipients = [e.strip() for e in alert_email.split(",") if e.strip()]
            if not recipients:
                print("⚠️ No valid email recipients found inside ALERT_EMAIL.")
            else:
                print(f"📧 Sending Unified Daily Watchlist Email to: {', '.join(recipients)}...")
                msg = MIMEMultipart("mixed")
                msg["Subject"] = f"🏆 Daily Monit Multibagger Watchlist — {today_str}"
                msg["From"] = gmail_user
                msg["To"] = ", ".join(recipients)
                
                msg.attach(MIMEText(html_text, "html"))
                
                # Attach compiled Excel workbook
                with open(str(excel_path), "rb") as f:
                    part = MIMEApplication(f.read(), Name=filename)
                    part['Content-Disposition'] = f'attachment; filename="{filename}"'
                    msg.attach(part)
                    
                with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
                    s.login(gmail_user, gmail_app_pass)
                    s.sendmail(gmail_user, recipients, msg.as_string())
                print("✅ Unified Email sent successfully.")
        except Exception as e:
            print(f"❌ Email delivery failed: {e}")
    else:
        print("ℹ️ Email credentials not configured in environment — skipping email alert.")
        
    # 5. Send Unified Telegram Messages
    if tg_token and tg_chat:
        tg_chat_ids = [c.strip() for c in str(tg_chat).split(",") if c.strip()]
        if not tg_chat_ids:
            print("⚠️ No valid Telegram Chat IDs found inside TELEGRAM_CHAT_ID.")
        else:
            print(f"📨 Sending Workbook & Alerts to {len(tg_chat_ids)} Telegram chat(s)...")
            doc_url = f"https://api.telegram.org/bot{tg_token}/sendDocument"
            msg_url = f"https://api.telegram.org/bot{tg_token}/sendMessage"
            
            for chat_id in tg_chat_ids:
                try:
                    # A. Send Workbook Document first
                    with open(str(excel_path), "rb") as f:
                        files = {"document": f}
                        data = {
                            "chat_id": chat_id,
                            "caption": f"{warning_tg}🏆 *Daily Monit Multibagger Watchlist — {today_str}*\n\n" \
                                       f"• Scored & Ranked: {universe_len} companies\n" \
                                       f"• Database: Updated & Synced\n\n" \
                                       f"Your daily watchlist workbook is attached above! 📈",
                            "parse_mode": "Markdown"
                        }
                        r = requests.post(doc_url, data=data, files=files, timeout=30)
                        r.raise_for_status()
                    print(f"   ✅ Workbook successfully delivered to chat: {chat_id}")
                    
                    # B. Send screened signals list and confluence/emerging report immediately after in the same chat
                    if has_signals or CONFLUENCE_EMERGING_REPORT:
                        chunks = [plain_text[i:i+4000] for i in range(0, len(plain_text), 4000)]
                        for chunk in chunks:
                            r = requests.post(
                                msg_url,
                                json={"chat_id": chat_id, "text": chunk, "parse_mode": ""},
                                timeout=15
                            )
                            r.raise_for_status()
                        print(f"   ✅ Watchlist details and report sent to chat: {chat_id}")
                except Exception as e:
                    print(f"   ❌ Telegram delivery failed for chat {chat_id}: {e}")
    else:
        print("ℹ️ Telegram credentials not configured in environment — skipping telegram alert.")

    # 6. Send Twilio WhatsApp Alert
    twilio_sid = os.environ.get("TWILIO_SID", "")
    twilio_token = os.environ.get("TWILIO_TOKEN", "")
    twilio_from = os.environ.get("TWILIO_WHATSAPP_FROM", "whatsapp:+14155238886")
    twilio_to = os.environ.get("TWILIO_WHATSAPP_TO", "")

    if twilio_sid and twilio_token and twilio_to:
        print("📨 Sending WhatsApp alert via Twilio...")
        try:
            from twilio.rest import Client
            client = Client(twilio_sid, twilio_token)
            wa_text = f"{warning_wa}🏆 *Daily Monit Watchlist — {today_str}*\n\n" \
                      f"• Scored & Ranked: {universe_len} companies\n" \
                      f"• Database: Updated & Synced\n\n"
            if has_signals:
                n_cross = len(signals_df[signals_df["entry"] == "EMA Crossover"])
                n_new   = len(signals_df[signals_df["entry"] == "52W Breakout"])
                n_run   = len(signals_df[signals_df["entry"] == "ATH Momentum"])
                wa_text += f"📊 *Signals Today:*\n" \
                           f"🟢 EMA Crossover: {n_cross}\n" \
                           f"🟡 52W Breakout: {n_new}\n" \
                           f"🔵 ATH Momentum: {n_run}\n\n" \
                           f"Your premium daily Excel workbook has been delivered to your Email and Telegram! 🚀"
            else:
                wa_text += "ℹ️ No active screener signals were generated today.\n\n" \
                           "Your premium daily Excel workbook has been delivered to your Email and Telegram!"
            
            client.messages.create(
                body=wa_text,
                from_=twilio_from,
                to=twilio_to
            )
            print("✅ WhatsApp alert sent successfully.")
        except Exception as e:
            print(f"❌ WhatsApp delivery failed: {e}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a Chartink-fed SOIC ranking workbook.")
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()

    output_path = args.output or DEFAULT_OUTPUT / f"monit_chartink_ranking_{date.today()}.xlsx"
    universe = fetch_chartink_universe()
    path = build_workbook(universe, args.template, output_path)
    print(f"Saved {path} with {len(universe)} Chartink companies")
    
    # Send daily cloud alerts (Email + Telegram) with the Excel workbook attached!
    send_cloud_alerts(path, len(universe))


if __name__ == "__main__":
    main()
